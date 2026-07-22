import json, ast, openpyxl, random, uuid
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, update_session_auth_hash, logout
from .models import (
    TipoSocio, Socio, CuentaBancaria, MetodoPago, Prestamo, Pago, 
    Bingo, Ahorro, Jugador, PartidaBingo, Carton, CartonPartidaBingo, 
    PlataformaJuego, SesionJuego, Regalo, AporteSemanal, ConfiguracionWeb, UnidadMonetaria, MensajeChat,
    validar_cedula_ecuatoriana, TarjetaRecarga, TransaccionRecarga, ValoracionSistema
)
from .services import  generar_lote_cartones, actualizar_avatar_perfil, validar_carton_hibrido
from django.contrib import messages
from django.contrib.auth.models import User
from django.db.models import Sum, Q, Avg
from django.db.models.deletion import ProtectedError
from django.contrib.auth.decorators import login_required
from datetime import datetime, date, timedelta
from django.db import transaction
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from channels.layers import get_channel_layer
from asgiref.sync import async_to_sync
from django.template.loader import get_template
from xhtml2pdf import pisa
from openpyxl.styles import Font, Alignment, PatternFill
from decimal import Decimal
from django.core.exceptions import ValidationError
from django.urls import reverse

# ===============================================================================================================================================
#                                                                   NOTA:
#                                               Seccion "COMUNES" empieza desde la linea 39
#                                               Seccion "CUENTAS" empieza desde la linea 
#                                               Seccion "LOGICA FINANCIERA" empieza desde la linea 
#                                               Seccion "ADMINISTRADOR" empieza desde la linea 
#                                               Seccion "PARTIDA" empieza desde la linea 
# ===============================================================================================================================================

# ===============================================================================================================================================
# 1. COMUNES (Páginas públicas y base)
# ===============================================================================================================================================
def inicio(request):
    """
    Vista principal de la página de inicio.
    Gestiona el estado del usuario (jugador o socio), calcula los pozos dinámicos 
    de los bingos activos y recopila las valoraciones del sistema.
    """
    # ==========================================================================================
    # 1. INICIALIZACIÓN DE FLAGS DE SESIÓN
    # ==========================================================================================
    preguntar_jugador = request.session.pop('preguntar_jugador', False)
    es_jugador = False
    es_socio = False
    mostrar_promo_socio = False
    # ==========================================================================================

    # ==========================================================================================
    # 2. VALIDACIÓN DE PERFIL DE USUARIO LOGUEADO
    # ==========================================================================================
    if request.user.is_authenticated and not request.user.is_staff:
        # Buscamos si existe un perfil de Jugador o Socio vinculado al username (cédula)
        jugador = Jugador.objects.filter(cedulaidentidadjugador=request.user.username).first()
        socio = Socio.objects.filter(cisocio=request.user.username).first()
        if jugador:
            es_jugador = True
            # Si el jugador aún no es socio, disparamos la promoción de ascenso una sola vez por sesión
            if not jugador.idsocio and not request.session.get('promo_socio_visto', False):
                mostrar_promo_socio = True
                request.session['promo_socio_visto'] = True

        if socio:
            es_socio = True
            # Validamos en sesión si el socio se encuentra en estado 'Activo'
            request.session['es_socio_activo'] = True if socio.estadosocio == 'Activo' else False
    # ==========================================================================================

    # ==========================================================================================
    # 3. CARGA DE CONFIGURACIÓN Y BINGOS ACTIVOS
    # ==========================================================================================
    config_web = ConfiguracionWeb.objects.first()
    # Traemos los bingos programados o en curso, optimizando las relaciones de divisas
    bingos_activos = Bingo.objects.filter(
        estadobingo__in=['Programado', 'En Curso']
    ).select_related('idunidad_venta', 'idunidad_premio').order_by('fechaprogramadabingo')  
    ahora = timezone.now() 
    # ==========================================================================================

    # ==========================================================================================
    # 4. LOGÍSTICA DE SALAS Y POZOS DINÁMICOS
    # ==========================================================================================
    for b in bingos_activos:
        # A) Control de apertura de sala (Se abre 30 minutos antes de la hora programada si hay partida activa)
        if b.fechaprogramadabingo:
            hora_apertura = b.fechaprogramadabingo - timedelta(minutes=30)
            partida_activa = PartidaBingo.objects.filter(idbingo=b,estadopartida__in=['Programada', 'En Juego']).order_by('idpartidabingo').first()
            if ahora >= hora_apertura and partida_activa:
                b.sala_abierta = True
                b.id_partida_a_entrar = partida_activa.idpartidabingo
            else:
                b.sala_abierta = False
        else:
            b.sala_abierta = False
        # B) Cálculo del Pozo Dinámico Multidivisa (Regla del 45%)
        vendidos = CartonPartidaBingo.objects.filter(idpartida__idbingo=b).values('idcarton').distinct().count()
        tasa_venta = float(b.idunidad_venta.tasaconversionmoneda)
        tasa_premio = float(b.idunidad_premio.tasaconversionmoneda)
        ingreso_en_dolares = float(vendidos * b.preciocarton) * tasa_venta
        fondo_pozo_dolares = ingreso_en_dolares * 0.45 # El 45% de la recaudación alimenta el pozo
        premio_base_dolares = float(b.premiomayor) * tasa_premio
        
        # Comparamos si el fondo dinámico supera al premio base inicial
        if fondo_pozo_dolares > premio_base_dolares:
            b.pozo_dinamico_actual = fondo_pozo_dolares / tasa_premio
        else:
            b.pozo_dinamico_actual = float(b.premiomayor)
    # ==========================================================================================

    # ==========================================================================================
    # 5. SISTEMA DE VALORACIONES Y RESEÑAS
    # ==========================================================================================
    valoraciones = ValoracionSistema.objects.all().order_by('-fecha')
    promedio = valoraciones.aggregate(promedio=Avg('puntuacion'))['promedio'] or 0
    total_valoraciones = valoraciones.count()

    mi_valoracion = None
    if request.user.is_authenticated:
        mi_valoracion = valoraciones.filter(usuario=request.user).first()
    # ==========================================================================================

    # ==========================================================================================
    # 6. EMPAQUETADO DE CONTEXTO Y RENDERIZADO
    # ==========================================================================================
    contexto = {
        'preguntar_jugador': preguntar_jugador,
        'es_jugador': es_jugador,
        'es_socio': es_socio,
        'config_web': config_web,
        'bingos_activos': bingos_activos,
        'mostrar_promo_socio': mostrar_promo_socio,
        'valoraciones': valoraciones,
        'promedio_valoracion': promedio,
        'total_valoraciones': total_valoraciones,
        'mi_valoracion': mi_valoracion,
    }
    return render(request, 'comunes/inicio.html', contexto)
    # ==========================================================================================

def bingo_publico(request):
    """
    Vista pública para mostrar la cartelera de eventos de bingo (activos y pasados).
    Calcula dinámicamente la apertura de las salas de espera según la hora y
    actualiza los pozos acumulados en función de los cartones vendidos (multidivisa).
    """
    # ==========================================================================================
    # 1. FILTRADO DE BINGOS (ACTIVOS E HISTORIAL)
    # ==========================================================================================
    # Eventos vigentes listos para venta o juego
    bingos_activos = Bingo.objects.filter(
        estadobingo__in=['Programado', 'En Curso']
    ).order_by('fechaprogramadabingo')
    # Eventos concluidos o cancelados (modo historial)
    bingos_pasados = Bingo.objects.filter(
        estadobingo__in=['Finalizado', 'Cancelado']
    ).order_by('-fechaprogramadabingo')
    ahora = timezone.now() 
    # ==========================================================================================

    # ==========================================================================================
    # 2. PROCESAMIENTO Y LÓGICA DE BINGOS ACTIVOS
    # ==========================================================================================
    for b in bingos_activos:
        # A) Control de acceso a la sala de espera (Habilita el ingreso 30 minutos antes)
        if b.fechaprogramadabingo:
            hora_apertura = b.fechaprogramadabingo - timedelta(minutes=30)
            partida_activa = PartidaBingo.objects.filter(
                idbingo=b,
                estadopartida__in=['Programada', 'En Juego']
            ).order_by('idpartidabingo').first()
            
            if ahora >= hora_apertura and partida_activa:
                b.sala_abierta = True
                b.id_partida_a_entrar = partida_activa.idpartidabingo
            else:
                b.sala_abierta = False
        else:
            b.sala_abierta = False
        # B) Cálculo del Pozo Dinámico Multidivisa (45% de la recaudación total)
        vendidos = CartonPartidaBingo.objects.filter(idpartida__idbingo=b).values('idcarton').distinct().count()
        tasa_venta = float(b.idunidad_venta.tasaconversionmoneda)
        tasa_premio = float(b.idunidad_premio.tasaconversionmoneda)
        ingreso_en_dolares = float(vendidos * b.preciocarton) * tasa_venta
        fondo_pozo_dolares = ingreso_en_dolares * 0.45
        premio_base_dolares = float(b.premiomayor) * tasa_premio
        # Evalúa si el pozo dinámico supera al premio base inicial establecido
        if fondo_pozo_dolares > premio_base_dolares:
            b.pozo_dinamico_actual = fondo_pozo_dolares / tasa_premio
        else:
            b.pozo_dinamico_actual = float(b.premiomayor)
    # ==========================================================================================
    
    # ==========================================================================================
    # 3. PROCESAMIENTO DE BINGOS PASADOS (HISTORIAL)
    # ==========================================================================================
    # Se recalcula el valor final que tuvo el pozo para mantener consistencia en la vista
    for b in bingos_pasados:
        vendidos = CartonPartidaBingo.objects.filter(idpartida__idbingo=b).values('idcarton').distinct().count()
        tasa_venta = float(b.idunidad_venta.tasaconversionmoneda)
        tasa_premio = float(b.idunidad_premio.tasaconversionmoneda)
        ingreso_en_dolares = float(vendidos * b.preciocarton) * tasa_venta
        fondo_pozo_dolares = ingreso_en_dolares * 0.45
        premio_base_dolares = float(b.premiomayor) * tasa_premio     
        if fondo_pozo_dolares > premio_base_dolares:
            b.pozo_dinamico_actual = fondo_pozo_dolares / tasa_premio
        else:
            b.pozo_dinamico_actual = float(b.premiomayor)
    # ==========================================================================================

    # ==========================================================================================
    # 4. EMPAQUETADO Y RENDERIZADO DE LA VISTA
    # ==========================================================================================
    contexto = {
        'bingos_activos': bingos_activos,
        'bingos_pasados': bingos_pasados,
        'unidad_monetaria': UnidadMonetaria.objects.first()
    }
    return render(request, 'comunes/bingo.html', contexto)
    # ==========================================================================================

@login_required
def agregar_valoracion(request):    
    """
    Vista protegida para que un usuario autenticado pueda registrar 
    o actualizar su valoración y comentario sobre la plataforma.
    Valida que la puntuación se encuentre en el rango permitido (0.0 a 5.0 estrellas).
    """
    # ==========================================================================================
    # 1. PROCESAMIENTO DE LA RESEÑA VÍA POST
    # ==========================================================================================
    if request.method == 'POST':
        estrellas = request.POST.get('estrellas')
        comentario = request.POST.get('comentario')
        try:
            # Convertimos el valor recibido a Decimal para validación matemática exacta
            puntuacion = Decimal(estrellas)
            # Verificamos el rango válido de estrellas
            if Decimal('0.0') <= puntuacion <= Decimal('5.0'):
                # Actualiza la reseña existente del usuario o crea una nueva si no existe
                ValoracionSistema.objects.update_or_create(
                    usuario=request.user,
                    defaults={'puntuacion': puntuacion, 'comentario': comentario}
                )
                messages.success(request, "¡Tu valoración ha sido guardada exitosamente!")
            else:
                messages.error(request, "La valoración debe estar entre 0 y 5 estrellas.")
        except Exception as e:
            messages.error(request, f"Error al guardar la valoración: {e}")
    # ==========================================================================================

    # ==========================================================================================
    # 2. REDIRECCIÓN A LA LANDING PAGE (ANCLAJE)
    # ==========================================================================================
    # Redirige de vuelta al inicio posicionando la vista directamente en la sección de reseñas
    return redirect(f"{reverse('inicio')}#seccion-contacto-resenas")
    # ==========================================================================================
# ===============================================================================================================================================

# ===============================================================================================================================================
# 2. CUENTAS (Autenticación y Perfiles)
# ===============================================================================================================================================
def seleccion_registro(request): return render(request, 'cuentas/seleccion_registro.html')

def registro_socio(request):
    """
    Vista encargada del autorregistro de nuevos socios en la cooperativa.
    Valida la integridad de los datos (cédula de 10 dígitos, algoritmo ecuatoriano,
    unicidad de correo y cédula) y crea tanto el usuario base de Django como el perfil
    de socio con estado 'Pendiente' de aprobación por el administrador.
    """
    if request.method == 'POST':
        # ==========================================================================================
        # 1. CAPTURA Y LIMPIEZA DE DATOS DEL FORMULARIO
        # ==========================================================================================
        primer_nombre = request.POST.get('primer_nombre')
        segundo_nombre = request.POST.get('segundo_nombre')
        primer_apellido = request.POST.get('primer_apellido')
        segundo_apellido = request.POST.get('segundo_apellido')
        cedula = request.POST.get('cedula')
        fecha_nacimiento_str = request.POST.get('fecha_nacimiento')
        nacionalidad = request.POST.get('nacionalidad', 'Ecuatoriana')
        telefono_personal = request.POST.get('telefono_personal') 
        direccion = request.POST.get('direccion')
        telefonofijo = request.POST.get('telefonofijo')
        direcciontrabajo = request.POST.get('direcciontrabajo')
        sexo = request.POST.get('sexo')
        email = request.POST.get('email')
        password = request.POST.get('password')
        # ==========================================================================================
        
        # ==========================================================================================
        # 2. ESCUDO DE INTEGRIDAD Y LONGITUD (10 DÍGITOS)
        # ==========================================================================================
        if not cedula.isdigit() or len(cedula) != 10:
            messages.error(request, "Error de seguridad: La cédula debe tener exactamente 10 dígitos numéricos.")
            return redirect('registro_socio')
            
        if not telefono_personal.isdigit() or len(telefono_personal) != 10:
            messages.error(request, "Error de seguridad: El teléfono debe tener exactamente 10 dígitos numéricos.")
            return redirect('registro_socio')
        # ==========================================================================================

        # ==========================================================================================
        # 3. VALIDACIÓN MATEMÁTICA DE CÉDULA ECUATORIANA
        # ==========================================================================================
        if nacionalidad == 'Ecuatoriana':
            try:
                validar_cedula_ecuatoriana(cedula)
            except ValidationError as e:
                messages.error(request, f"Error de verificación: {e.message}")
                return render(request, 'cuentas/registro_socio.html')
        # ==========================================================================================
            
        # ==========================================================================================
        # 4. VERIFICACIÓN DE UNICIDAD (DUPLICADOS)
        # ==========================================================================================
        if User.objects.filter(username=cedula).exists():
            messages.error(request, "Esta cédula ya está registrada.")
            return redirect('registro_socio')
            
        if User.objects.filter(email=email).exists():
            messages.error(request, "Este correo electrónico ya está registrado.")
            return redirect('registro_socio')
        # ==========================================================================================

        # ==========================================================================================
        # 5. VALIDACIÓN LÓGICA DE FECHA DE NACIMIENTO
        # ==========================================================================================
        try:
            fecha_nac = datetime.strptime(fecha_nacimiento_str, '%Y-%m-%d').date()
            if fecha_nac > date.today():
                messages.error(request, "La fecha de nacimiento no puede ser en el futuro.")
                return redirect('registro_socio')
        except ValueError:
            messages.error(request, "Formato de fecha inválido.")
            return redirect('registro_socio')
        # ==========================================================================================

        # ==========================================================================================
        # 6. CREACIÓN SEGURA DE USUARIO Y PERFIL DE SOCIO
        # ==========================================================================================
        try:
            # Creamos las credenciales base en el modelo auth de Django
            user = User.objects.create_user(
                username=cedula, 
                email=email, 
                password=password, 
                first_name=primer_nombre, 
                last_name=primer_apellido)
            # Verificamos que existan tipos de socio configurados
            tipo_base = TipoSocio.objects.first()
            if not tipo_base:
                user.delete() 
                messages.error(request, "Error crítico: No hay 'Tipos de Socio'.")
                return redirect('registro_socio')
            # Creamos el registro del Socio asociado, fijando el estado inicial como 'Pendiente'
            Socio.objects.create(
                idtiposocio=tipo_base, 
                primernombresocio=primer_nombre, 
                segundonombresocio=segundo_nombre,
                primerapellidosocio=primer_apellido, 
                segundoapellidosocio=segundo_apellido, 
                cisocio=cedula,
                fechanacimientosocio=fecha_nac, 
                telefonopersonalsocio=telefono_personal,
                direcciondomiciliosocio=direccion, 
                sexosocio=sexo, 
                estadosocio='Pendiente', # Requiere validación del administrador
                nacionalidad=nacionalidad,
                telefonotrabajosocio=telefonofijo, 
                direcciontrabajosocio=direcciontrabajo,
                correosocio=email
            )
            messages.success(request, "¡Registro completado! Tu solicitud ha sido enviada. Por favor, espera a que un administrador valide tu cuenta para poder iniciar sesión.")
            return redirect('inicio')
        except Exception as e:
            # Control de errores y rollback: si algo falla al crear el socio, eliminamos el usuario huérfano
            if 'user' in locals() and user.id: user.delete() 
            messages.error(request, f"Error en el formulario: {str(e)}")
            return redirect('registro_socio')
        # ==========================================================================================
    
    # Renderizado inicial del formulario de registro en caso de método GET
    return render(request, 'cuentas/registro_socio.html')
    # ==========================================================================================

def registro_jugador(request):
    """
    Vista dual para el registro o activación de perfiles de juego.
    1. Si el usuario ya está autenticado (es Socio), vincula y crea su perfil de jugador automáticamente.
    2. Si es un usuario externo, valida su cédula, crea sus credenciales en el sistema, 
       inicia sesión de forma automática y lo redirige a la sala de juegos.
    """
    if request.method == 'POST':
        # ==========================================================================================
        # 1. CAPTURA DE DATOS BÁSICOS
        # ==========================================================================================
        alias = request.POST.get('aliasjugador')
        nacionalidad = request.POST.get('nacionalidad', 'Ecuatoriana') 
        # ==========================================================================================

        # ==========================================================================================
        # 2. ESCENARIO A: SOCIO YA AUTENTICADO ACTIVANDO JUEGO
        # ==========================================================================================
        if request.user.is_authenticated:
            try:
                # Buscamos al socio asociado a la cuenta actual del usuario
                socio_vinculado = Socio.objects.get(cisocio=request.user.username)
                # Creamos el perfil de jugador heredando directamente los datos del Socio
                Jugador.objects.create(
                    idsocio=socio_vinculado, 
                    aliasjugador=alias, 
                    nombresjugador=socio_vinculado.primernombresocio, 
                    cedulaidentidadjugador=socio_vinculado.cisocio, 
                    correojugador=request.user.email,
                    nacionalidad=nacionalidad
                )
                # Actualizamos las variables de sesión y notificamos el éxito
                request.session['user_nombre'] = alias
                messages.success(request, f"¡Perfil de juego activado como '{alias}'!")
                return redirect('inicio')
            except Exception:
                messages.error(request, "Error al vincular el perfil de juego.")
        # ==========================================================================================

        # ==========================================================================================
        # 3. ESCENARIO B: REGISTRO DE JUGADOR EXTERNO
        # ==========================================================================================
        else:
            nombres, apellidos = request.POST.get('nombresjugador'), request.POST.get('apellidosjugador')
            cedula, correo, password = request.POST.get('cedula'), request.POST.get('correojugador'), request.POST.get('password')
            # A) Validación estricta de longitud y formato numérico de la cédula
            if cedula and (not cedula.isdigit() or len(cedula) != 10):
                messages.error(request, "Error de seguridad: La cédula debe tener exactamente 10 dígitos numéricos.")
                return redirect('registro_jugador')
            # B) Validación algorítmica de la cédula ecuatoriana (si aplica)
            if nacionalidad == 'Ecuatoriana':
                try:
                   validar_cedula_ecuatoriana(cedula)
                except ValidationError as e:
                    messages.error(request, f"Error de verificación: {e.message}")
                    return redirect('registro_jugador')
            # C) Verificación de que la cédula no esté previamente registrada en el sistema
            if User.objects.filter(username=cedula).exists():
                messages.error(request, "Cédula ya registrada.")
                return redirect('registro_jugador')
            # D) Transacción segura para crear usuario de autenticación y perfil de jugador
            try:
                # Creamos el usuario base de Django (usando la cédula como username)
                user = User.objects.create_user(
                    username=cedula, 
                    email=correo, 
                    password=password, 
                    first_name=nombres, 
                    last_name=apellidos)  
                # Creamos el registro en el modelo Jugador
                Jugador.objects.create(
                    aliasjugador=alias, 
                    nombresjugador=nombres, 
                    apellidosjugador=apellidos, 
                    cedulaidentidadjugador=cedula, 
                    correojugador=correo,
                    nacionalidad=nacionalidad 
                )
                # Autenticamos al usuario de inmediato y establecemos la sesión
                login(request, user)
                request.session['user_nombre'] = alias
                messages.success(request, f"¡Bienvenido a la sala de juegos, {alias}!")
                return redirect('inicio')
            except Exception as e:
                # Rollback de seguridad: si ocurre un fallo al crear el jugador, eliminamos el usuario de Django creado
                if 'user' in locals() and user.id: user.delete()
                messages.error(request, f"Error: {str(e)}")
                return redirect('registro_jugador')
        # ==========================================================================================    

    # Renderizado inicial del formulario de registro de jugador (método GET)            
    return render(request, 'cuentas/registro_jugador.html')
    # ==========================================================================================

def inicio_sesion(request):
    """
    Vista encargada de gestionar el inicio de sesión de los usuarios en la plataforma.
    Realiza validaciones de credenciales (usuario o correo), verifica si la cuenta 
    se encuentra activa y aplica restricciones especiales para socios en estado 'Pendiente'.
    Configura las variables de sesión globales (nombre, avatar y estado de socio).
    """
    # ==========================================================================================
    # 1. REDIRECCIÓN SI YA ESTÁ AUTENTICADO
    # ==========================================================================================
    if request.user.is_authenticated: return redirect('dashboard' if request.user.is_staff else 'inicio')
    # ==========================================================================================

    # ==========================================================================================
    # 2. PROCESAMIENTO DE CREDENCIALES (VÍA POST)
    # ==========================================================================================
    if request.method == 'POST':
        identificador = request.POST.get('identificador')
        password = request.POST.get('password')
        # Buscamos al usuario de Django permitiendo iniciar sesión tanto por username como por email
        user = User.objects.filter(Q(username=identificador) | Q(email=identificador)).first()
        if user and user.check_password(password):
            # A) Verificamos que la cuenta base de Django no esté desactivada o suspendida
            if not user.is_active:
                messages.error(request, "Esta cuenta ha sido desactivada o suspendida del sistema.")
                return redirect('inicio_sesion')
            # B) Buscamos perfiles vinculados de Socio y Jugador utilizando la cédula (username)
            socio = Socio.objects.filter(cisocio=user.username).first()
            jugador = Jugador.objects.filter(cedulaidentidadjugador=user.username).first()
            # C) Escudo de seguridad para solicitudes de Socio en estado 'Pendiente'
            if socio and socio.estadosocio == 'Pendiente':
                if not jugador:
                    # Si es un socio nuevo sin perfil de juego previo, se bloquea el acceso hasta que sea aprobado
                    messages.warning(request, "Tu cuenta de socio aún está en revisión. Un administrador debe validarla antes de que puedas entrar.")
                    return redirect('inicio_sesion')
                else:
                    # Si ya posee una cuenta de jugador previa, se le permite el acceso pero se le notifica el estado pendiente
                    messages.info(request, "Aviso: Tu solicitud para ascender a Socio aún está en revisión por el administrador. Mientras tanto, puedes usar tu cuenta de jugador.")
            # ==========================================================================================
            # 3. AUTENTICACIÓN Y CONFIGURACIÓN DE SESIÓN
            # ==========================================================================================
            login(request, user)
            jugador = Jugador.objects.filter(cedulaidentidadjugador=user.username).first()
            # Determinamos el nombre y avatar que se mostrarán en la interfaz de navegación
            nombre_mostrar = user.first_name
            avatar_url = None
            if jugador:
                nombre_mostrar = jugador.aliasjugador or user.first_name
                if jugador.avatarjugador: avatar_url = jugador.avatarjugador.url
            if socio and not avatar_url:
                if socio.fotosocio: avatar_url = socio.fotosocio.url
            # Almacenamos los datos clave en las variables de sesión de Django
            request.session['user_nombre'] = nombre_mostrar
            request.session['avatar_url'] = avatar_url
            request.session['es_socio_activo'] = True if (socio and socio.estadosocio == 'Activo') else False
            messages.success(request, f"¡Bienvenido de vuelta, {nombre_mostrar}!")
            # Redirección inteligente: Panel de control si es admin, de lo contrario a la vista de inicio
            return redirect('dashboard' if user.is_staff else 'inicio')
            # ==========================================================================================
        else:
            messages.error(request, "Credenciales incorrectas. Verifica tu usuario/cédula/correo y contraseña.")
        # ==========================================================================================
    # Renderizado inicial de la plantilla de inicio de sesión (método GET)
    return render(request, 'cuentas/inicio_sesion.html')
    # ==========================================================================================

def cerrar_sesion(request):
    logout(request)
    return redirect('inicio')

@login_required
def perfil(request):
    """
    Vista protegida para la gestión integral del perfil del usuario (Socio y/o Jugador).
    Maneja peticiones POST para:
    1. Actualizar datos básicos de contacto.
    2. Actualizar la foto de perfil (avatar).
    3. Cambiar de contraseña de forma segura.
    4. Solicitar el ascenso de Jugador a Socio (creando un registro en estado 'Pendiente').
    Y mediante peticiones GET, recopila y muestra los historiales de compras, préstamos y ahorros.
    """
    user = request.user
    socio = Socio.objects.filter(cisocio=user.username).first()
    jugador = Jugador.objects.filter(cedulaidentidadjugador=user.username).first()
    # ==========================================================================================
    # 1. PROCESAMIENTO DE ACCIONES VÍA POST
    # ==========================================================================================
    if request.method == 'POST':
        action = request.POST.get('action')
        try:
            # A) Actualización de datos básicos de contacto
            if action == 'actualizar_datos':
                nuevo_correo = request.POST.get('correo')
                if nuevo_correo:
                    user.email = nuevo_correo
                    user.save()
                
                if socio:
                    socio.telefonopersonalsocio = request.POST.get('telefono', socio.telefonopersonalsocio)
                    socio.save()
                    
                if jugador:
                    jugador.aliasjugador = request.POST.get('alias', jugador.aliasjugador)
                    jugador.correojugador = nuevo_correo
                    jugador.save()
                    request.session['user_nombre'] = jugador.aliasjugador
                    
                messages.success(request, "Tus datos de contacto han sido actualizados.")
            # B) Actualización de foto de perfil (Avatar)
            elif action == 'actualizar_avatar':
                nueva_foto = request.FILES.get('avatar')
                if nueva_foto:
                    # Llama al servicio encargado de guardar la nueva imagen y limpiar el servidor de la anterior
                    actualizar_avatar_perfil(request, socio, jugador, nueva_foto)
                    messages.success(request, "¡Tu foto de perfil luce genial! (Servidor optimizado y limpio).")
            # C) Cambio de contraseña de seguridad
            elif action == 'actualizar_password':
                actual = request.POST.get('password_actual')
                nueva = request.POST.get('password_nueva')
                
                if user.check_password(actual):
                    user.set_password(nueva)
                    user.save()
                    # Mantiene la sesión activa para evitar que el usuario sea desconectado tras el cambio
                    update_session_auth_hash(request, user)
                    messages.success(request, "Tu contraseña ha sido cambiada de forma segura.")
                else:
                    messages.error(request, "La contraseña actual no coincide. No se guardaron los cambios.")
            # D) Solicitud de ascenso a Socio desde un perfil de Jugador
            elif action == 'ascender_socio':
                cedula = request.POST.get('cedula')
                primer_nombre = request.POST.get('primer_nombre')
                segundo_nombre = request.POST.get('segundo_nombre', '')
                primer_apellido = request.POST.get('primer_apellido')
                segundo_apellido = request.POST.get('segundo_apellido')
                telefono = request.POST.get('telefono')
                direccion = request.POST.get('direccion')
                fecha_nacimiento_str = request.POST.get('fecha_nacimiento')
                sexo = request.POST.get('sexo')
                correo_elegido = request.POST.get('correo') 
                telefonofijo = request.POST.get('telefonofijo')
                direcciontrabajo = request.POST.get('direcciontrabajo')
                
                try:
                    fecha_nac = datetime.strptime(fecha_nacimiento_str, '%Y-%m-%d').date()
                    tipo_base = TipoSocio.objects.first()
                    # Creamos el nuevo registro de socio con estado 'Pendiente' de validación administrativa
                    nuevo_socio = Socio.objects.create(
                        idtiposocio=tipo_base,
                        primernombresocio=primer_nombre,
                        segundonombresocio=segundo_nombre,
                        primerapellidosocio=primer_apellido,
                        segundoapellidosocio=segundo_apellido,
                        cisocio=cedula,
                        fechanacimientosocio=fecha_nac,
                        telefonopersonalsocio=telefono,
                        telefonotrabajosocio=telefonofijo, 
                        direcciondomiciliosocio=direccion,
                        direcciontrabajosocio=direcciontrabajo,
                        sexosocio=sexo,
                        correosocio=correo_elegido,
                        estadosocio='Pendiente' 
                    )
                    # Actualizamos las credenciales base del usuario en Django Auth
                    user.username = cedula
                    user.first_name = primer_nombre
                    user.last_name = primer_apellido
                    user.email = correo_elegido 
                    user.save()
                    # Vinculamos el perfil de jugador existente con el nuevo socio creado
                    jugador.idsocio = nuevo_socio
                    jugador.correojugador = correo_elegido
                    jugador.save()
                    
                    messages.success(request, "¡Solicitud enviada exitosamente! Tu perfil de Socio está en estado 'Pendiente' hasta que el Administrador lo valide. Mientras tanto, puedes seguir jugando con normalidad.")
                except Exception as e:
                    messages.error(request, f"Error al procesar la solicitud de socio: {str(e)}")

        except Exception as e:
            messages.error(request, f"Error al actualizar el perfil: {str(e)}")

        return redirect('perfil')
    # ==========================================================================================

    # ==========================================================================================
    # 2. CARGA DE HISTORIALES PARA EL RENDER (GET)
    # ==========================================================================================
    historial_compras = []
    historial_prestamos = []
    historial_ahorros = []
    
    if jugador:
        historial_compras = CartonPartidaBingo.objects.filter(idjugador=jugador).select_related('idpartida', 'idcarton').order_by('-fechacompra')[:15]
    if socio:
        historial_prestamos = Prestamo.objects.filter(idsocio=socio).order_by('-fechasolicitud')
        historial_ahorros = Ahorro.objects.filter(idsocio=socio).order_by('-fechaahorro')[:15]
    # ==========================================================================================

    # ==========================================================================================
    # 3. EMPAQUETADO DE CONTEXTO Y RENDERIZADO
    # ==========================================================================================
    contexto = {
        'socio': socio,
        'jugador': jugador,
        'historial_compras': historial_compras,
        'historial_prestamos': historial_prestamos,
        'historial_ahorros': historial_ahorros,
    }
    return render(request, 'cuentas/perfil.html', contexto)
    # ==========================================================================================

@login_required
def activar_perfil_juego_socio(request):
    """
    Vista para que un Socio activo pueda activar instantáneamente su perfil 
    de jugador y billetera en la plataforma.
    Valida que el socio exista y se encuentre activo, evita duplicados, 
    genera un alias único automáticamente y configura las variables de sesión.
    """
    # ==========================================================================================
    # 1. LOCALIZACIÓN Y VALIDACIÓN DEL SOCIO
    # ==========================================================================================
    socio = Socio.objects.filter(cisocio=request.user.username).first()
    
    if not socio:
        messages.error(request, "No se encontró un perfil de Socio vinculado a esta cuenta.")
        return redirect('inicio')
        
    if socio.estadosocio != 'Activo':
        messages.warning(request, "Tu cuenta de Socio debe estar aprobada y activa para habilitar el perfil de juego.")
        return redirect('perfil')
    # ==========================================================================================
        
    # ==========================================================================================
    # 2. CONTROL ANTI-DUPLICADOS DE JUGADOR
    # ==========================================================================================
    jugador_existente = Jugador.objects.filter(cedulaidentidadjugador=socio.cisocio).exists()
    if jugador_existente:
        messages.info(request, "Tu perfil de juego ya se encuentra activo.")
        return redirect('perfil')
    # ==========================================================================================
        
    # ==========================================================================================
    # 3. CREACIÓN DEL PERFIL DE JUGADOR
    # ==========================================================================================
    try:
        # Generación automática de un alias único (evita colisiones si el nombre ya está tomado)
        alias_propuesto = socio.primernombresocio
        if Jugador.objects.filter(aliasjugador=alias_propuesto).exists():

            alias_propuesto = f"{socio.primernombresocio}{socio.cisocio[-4:]}"
        # Creación del registro heredando la información directamente del Socio validado
        nuevo_jugador = Jugador.objects.create(
            idsocio=socio,
            aliasjugador=alias_propuesto,
            cedulaidentidadjugador=socio.cisocio,
            correojugador=socio.correosocio or request.user.email,
            nacionalidad=socio.nacionalidad,
            nombresjugador=None, # Limpieza de redundancia; los datos se leen desde la entidad Socio
            apellidosjugador=None,
            saldocreditojugador=Decimal('0.00'),
            saldovirtualjugador=Decimal('0.00'),
            estadocuentajugador='Activo'
        )
        
        # ==========================================================================================
        # 4. CONFIGURACIÓN DE SESIÓN Y NOTIFICACIÓN
        # ==========================================================================================
        request.session['user_nombre'] = nuevo_jugador.aliasjugador
        if socio.fotosocio:
            request.session['avatar_url'] = socio.fotosocio.url
            
        messages.success(request, f"¡Billetera y perfil de juego activados al instante! Tu alias inicial es '{alias_propuesto}' (puedes cambiarlo cuando desees).")
        return redirect('perfil')
        # ==========================================================================================
    # ==========================================================================================
        
    except Exception as e:
        messages.error(request, f"Error crítico al activar el perfil de juego: {str(e)}")
        return redirect('perfil')
    # ==========================================================================================

@login_required
def mis_cartones(request):
    """
    Vista protegida para que el jugador consulte todos los cartones 
    que ha adquirido agrupados por su respectivo evento de Bingo.
    Evalúa de forma inteligente si los cartones pueden ser cambiados 
    según el estado actual de las rondas de juego.
    """
    ## ==========================================================================================
    # 1. IDENTIFICACIÓN Y VALIDACIÓN DEL JUGADOR
    # ==========================================================================================
    # Buscamos al jugador utilizando la cédula almacenada en el username del usuario autenticado
    jugador = Jugador.objects.filter(cedulaidentidadjugador=request.user.username).first()
    
    if not jugador:
        messages.warning(request, "Debes activar tu perfil de juego para ver tus cartones.")
        return redirect('inicio')
    # ==========================================================================================

    # ==========================================================================================
    # 2. CONSULTA OPTIMIZADA DE CARTONES ASIGNADOS
    # ==========================================================================================
    # Traemos las relaciones de cartones ordenadas del evento más reciente al más antiguo        
    cartones_jugador = CartonPartidaBingo.objects.filter(idjugador=jugador).select_related(
        'idcarton', 'idpartida', 'idpartida__idbingo'
    ).order_by('-idpartida__idbingo__fechaprogramadabingo')
    # ==========================================================================================
    
    # ==========================================================================================
    # 3. DEDUPLICACIÓN DE CARTONES POR BINGO
    # ==========================================================================================
    # Creamos un diccionario organizador para evitar mostrar cartones repetidos por cada ronda
    bingos_dict = {}
    for c in cartones_jugador:
        b_id = c.idpartida.idbingo.idbingo
        if b_id not in bingos_dict:
            bingos_dict[b_id] = {
                'bingo': c.idpartida.idbingo,
                'cartones_unicos': {} # Diccionario interno para filtrar IDs de cartones duplicados
            }
        # Almacenamos el cartón usando su ID como llave única
        carton_id = c.idcarton.idcarton
        if carton_id not in bingos_dict[b_id]['cartones_unicos']:
            bingos_dict[b_id]['cartones_unicos'][carton_id] = c
    # ==========================================================================================

    # ==========================================================================================
    # 4. EVALUACIÓN DE REGLAS PARA CAMBIO DE CARTÓN
    # ==========================================================================================
    bingos_agrupados = []
    for b_id, data in bingos_dict.items():
        # Verificamos si existe alguna ronda en curso o en proceso de verificación para este bingo
        rondas_en_juego = PartidaBingo.objects.filter(
            idbingo=data['bingo'],
            estadopartida__in=['En Juego', 'Verificando', 'Desempate']
        ).exists()
        # El cambio de cartón solo se permite si NO hay rondas activas y el bingo no ha concluido
        puede_cambiar = not rondas_en_juego and data['bingo'].estadobingo not in ['Finalizado', 'Cancelado']
        bingos_agrupados.append({
            'bingo': data['bingo'],
            'cartones': list(data['cartones_unicos'].values()),
            'puede_cambiar': puede_cambiar
        })
    # ==========================================================================================
        
    # ==========================================================================================
    # 5. EMPAQUETADO DE CONTEXTO Y RENDERIZADO
    # ========================================================================================== 
    contexto = {
        'bingos_agrupados': bingos_agrupados,
        'jugador': jugador
    }
    return render(request, 'cuentas/mis_cartones.html', contexto)
    # ==========================================================================================

@login_required
def descargar_cartones_pdf(request, id_bingo):
    if request.method == 'POST':
        jugador = Jugador.objects.filter(cedulaidentidadjugador=request.user.username).first()
        if not jugador:
            messages.error(request, "Perfil no encontrado.")
            return redirect('mis_cartones')

        cartones_ids = request.POST.getlist('cartones_seleccionados')
        if not cartones_ids:
            messages.warning(request, "No seleccionaste ningún cartón para imprimir.")
            return redirect('mis_cartones')

        bingo = get_object_or_404(Bingo, idbingo=id_bingo)
        cartones_asignados = CartonPartidaBingo.objects.filter(
            idjugador=jugador, 
            idpartida__idbingo=bingo,
            idcarton__in=cartones_ids
        ).select_related('idcarton')


        cartones_unicos = {}
        for asig in cartones_asignados:
            if asig.idcarton.idcarton not in cartones_unicos:
                matriz = asig.idcarton.matriznumeros
                if isinstance(matriz, str):
                    try: matriz = json.loads(matriz.replace("'", '"'))
                    except: continue
                
                if isinstance(matriz, dict):
                    try:
                        filas = []
                        for i in range(5):
                            filas.append([matriz['B'][i], matriz['I'][i], matriz['N'][i], matriz['G'][i], matriz['O'][i]])
                        
                        cartones_unicos[asig.idcarton.idcarton] = {
                            'codigo': asig.idcarton.codigocarton,
                            'filas': filas
                        }
                    except Exception as e: print(e)

        cartones_procesados = list(cartones_unicos.values())

        template = get_template('cuentas/cartones_pdf.html')
        context = {'bingo': bingo, 'jugador': jugador, 'cartones': cartones_procesados}
        html = template.render(context)

        response = HttpResponse(content_type='application/pdf')

        response['Content-Disposition'] = f'inline; filename="Mis_Cartones_{bingo.idbingo}_{jugador.aliasjugador}.pdf"'
        
        pisa_status = pisa.CreatePDF(html, dest=response)
        
        if pisa_status.err:
            return HttpResponse('Tuvimos errores generando tu documento PDF', status=500)
        return response
    
    return redirect('mis_cartones')

@login_required
def cambiar_carton_boveda(request):
    """Procesa el reemplazo del cartón validando reglas de negocio"""
    if request.method == 'POST':
        jugador = Jugador.objects.filter(cedulaidentidadjugador=request.user.username).first()
        id_bingo = request.POST.get('id_bingo')
        id_carton_viejo = request.POST.get('id_carton_viejo')
        modalidad = request.POST.get('modalidad') 
        id_carton_nuevo = request.POST.get('id_carton_nuevo') 

        bingo = get_object_or_404(Bingo, idbingo=id_bingo)

        # Regla 1: El evento no debe estar finalizado
        if bingo.estadobingo in ['Finalizado', 'Cancelado']:
            messages.error(request, "El evento ya concluyó, no puedes realizar cambios.")
            return redirect('mis_cartones')

        # Regla 2: Ninguna ronda debe estar en juego activo
        rondas_activas = PartidaBingo.objects.filter(
            idbingo=bingo, 
            estadopartida__in=['En Juego', 'Verificando', 'Desempate']
        ).exists()

        if rondas_activas:
            messages.error(request, "¡Acción bloqueada! Hay una ronda en juego en este instante.")
            return redirect('mis_cartones')

        # Regla 3: Filtrar SOLO las asignaciones de rondas futuras (Programada)
        asignaciones_pendientes = CartonPartidaBingo.objects.filter(
            idjugador=jugador, 
            idpartida__idbingo=bingo, 
            idcarton_id=id_carton_viejo,
            idpartida__estadopartida='Programada'
        )

        if not asignaciones_pendientes.exists():
            messages.error(request, "No tienes rondas futuras con este cartón para poder cambiarlo.")
            return redirect('mis_cartones')

        try:
            with transaction.atomic():
                nuevo_carton_obj = None

                if modalidad == 'rng':
                    lote = generar_lote_cartones(1) 
                    c_data = lote[0]
                    nuevo_carton_obj = Carton.objects.create(
                        codigocarton=c_data['codigo'], 
                        matriznumeros=c_data['matriz'], 
                        esmaestro=False
                    )
                
                elif modalidad == 'catalogo':
                    nuevo_carton_obj = Carton.objects.get(idcarton=id_carton_nuevo, esmaestro=True)
                    en_uso = CartonPartidaBingo.objects.filter(
                        idpartida__idbingo=bingo, 
                        idcarton=nuevo_carton_obj
                    ).exists()
                    
                    if en_uso:
                        messages.error(request, "¡Alguien más tomó ese cartón del catálogo! Intenta con otro.")
                        return redirect('mis_cartones')

                # FIX: Solo actualizamos las asignaciones de las rondas que aún no se juegan
                asignaciones_pendientes.update(idcarton=nuevo_carton_obj)
                
                messages.success(request, f"¡Cartón cambiado exitosamente! Tu nuevo código para las rondas restantes es {nuevo_carton_obj.codigocarton}.")

        except Exception as e:
            messages.error(request, f"Error al realizar el cambio: {str(e)}")

    return redirect('mis_cartones')
# ===============================================================================================================================================

# ===============================================================================================================================================
# 3. LOGICA FINANCIERA (Asociado a Cuentas)
# ===============================================================================================================================================
@login_required
def creditos(request):
    """
    Vista para que el Socio vea sus préstamos activos y solicite nuevos créditos.
    """
    socio = Socio.objects.filter(cisocio=request.user.username).first()
    
    # Validación de seguridad: Solo socios activos
    if not socio or socio.estadosocio not in ['Activo', 'Active']:
        messages.error(request, "Acceso denegado: Solo los socios activos pueden solicitar créditos.")
        return redirect('inicio')

    # Obtenemos los préstamos del socio ordenados por el más reciente
    mis_prestamos = Prestamo.objects.filter(idsocio=socio).order_by('-fechasolicitud')
    
    # Obtenemos garantes disponibles (excluyendo al solicitante)
    lista_socios = Socio.objects.filter(estadosocio='Activo').exclude(idsocio=socio.idsocio)

    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'solicitar_prestamo':
            # Nombres de los inputs del formulario HTML
            monto_solicitado = request.POST.get('montoprestamosolicitado')
            numerocuotas = request.POST.get('numerocuotas')
            fechavencimiento = request.POST.get('fechavencimiento')
            idgarante1 = request.POST.get('idgarante1')
            idgarante2 = request.POST.get('idgarante2')
            
            try:
                monto = float(monto_solicitado)
                cuotas = int(numerocuotas)
                
                # =========================================================
                # LÓGICA FINANCIERA: TASA FIJA DEL 10% (Regla de Vanessa)
                # =========================================================
                # Ignoramos cualquier input de tasa del usuario y forzamos el 10%
                tasa = Decimal('10.00') 
                
                if monto > 0 and cuotas > 0:
                    # Cálculo de intereses con precisión Decimal
                    interes_calculado = (Decimal(str(monto)) * tasa) / Decimal('100')
                    monto_total = Decimal(str(monto)) + interes_calculado
                    
                    garante1_obj = Socio.objects.filter(idsocio=idgarante1).first() if idgarante1 else None
                    garante2_obj = Socio.objects.filter(idsocio=idgarante2).first() if idgarante2 else None

                    # === MAGIA NUEVA: Leer método de desembolso ===
                    id_cuenta_destino = request.POST.get('idcuentadestino')
                    cuenta_obj = None
                    if id_cuenta_destino and id_cuenta_destino != 'EFECTIVO':
                        cuenta_obj = CuentaBancaria.objects.filter(idcuentabancaria=id_cuenta_destino).first()
                    # ==============================================
                    
                    # ⚠️ IMPORTANTE: Ajusta estos nombres si tu modelo usa 'montoprestamo' o 'saldovivoprestamo'
                    Prestamo.objects.create(
                        idsocio=socio,
                        idcuentadestino=cuenta_obj,
                        idgarante1=garante1_obj,
                        idgarante2=garante2_obj,
                        montoprestamosolicitado=monto, # Revisa si es montoprestamo
                        tasainteres=tasa,
                        numerocuotas=cuotas,
                        montototalpagar=monto_total,   # Revisa si es montototalapagar
                        saldopendiente=monto_total,    # Revisa si es saldovivoprestamo
                        fechasolicitud=timezone.now(),
                        fechavencimiento=fechavencimiento, 
                        estadoprestamo='Solicitado'
                    )
                    
                    messages.success(request, f"¡Tu solicitud por ${monto:.2f} ha sido enviada! Se aplicó la tasa oficial del 10%. Total a pagar: ${monto_total:.2f}.")
                else:
                    messages.error(request, "El monto y las cuotas deben ser valores mayores a 0.")
                    
            except Exception as e:
                messages.error(request, f"Ocurrió un error al procesar tu solicitud: {str(e)}")
                
            # Es mejor redirigir a la misma página de créditos para ver el cambio reflejado inmediatamente
            return redirect('creditos')

    mis_cuentas = CuentaBancaria.objects.filter(idsocio=socio, estadocuenta='Activo')

    contexto = {
        'socio': socio,
        'mis_prestamos': mis_prestamos,
        'lista_socios': lista_socios,
        'mis_cuentas': mis_cuentas
    }
    return render(request, 'negocio/creditos.html', contexto)

@login_required
def ahorro(request):
    """
    Vista para que el Socio vea su libreta de ahorros, reporte depósitos y solicite retiros.
    """
    socio = Socio.objects.filter(cisocio=request.user.username).first()
    
    # 1. Validación Estricta: Solo socios aprobados y activos
    if not socio or socio.estadosocio not in ['Activo', 'Active']:
        messages.warning(request, "Acceso denegado: Debes ser un Socio activo para acceder a la libreta de ahorros.")
        return redirect('inicio')

    historial_ahorros = Ahorro.objects.filter(idsocio=socio).order_by('-fechaahorro')
    total_ahorrado = historial_ahorros.filter(estadoahorro='Acreditado').aggregate(total=Sum('montoahorro'))['total'] or Decimal('0.00')

    if request.method == 'POST':
        action = request.POST.get('action')
        
        # ==============================================================
        # ACCIÓN 1: REPORTAR UN NUEVO DEPÓSITO
        # ==============================================================
        if action == 'registrar_ahorro':
            monto_ahorro = request.POST.get('monto_ahorro')
            id_metodo_pago = request.POST.get('idmetodopago')
            imagen_comprobante = request.FILES.get('comprobanteahorro')
            
            try:
                monto = float(monto_ahorro)
                bingo_vinculado = Bingo.objects.exclude(estadobingo__in=['Finalizado', 'Cancelado']).first()
                
                # Ya no creamos cuentas fantasma, buscamos el método oficial de la cooperativa
                metodo_obj = MetodoPago.objects.filter(idmetodopago=id_metodo_pago).first()
                
                if metodo_obj and monto > 0 and imagen_comprobante:
                    Ahorro.objects.create(
                        idsocio=socio,
                        idbingo=bingo_vinculado, 
                        idmetodopago=metodo_obj, # <-- Vinculamos la llave foránea
                        tipoahorro='Voluntario',
                        montoahorro=monto,
                        comprobanteahorro=imagen_comprobante,
                        fechaahorro=timezone.now(),
                        estadoahorro='Pendiente',
                        origenahorro='Directo'
                    )
                    messages.success(request, "¡Tu depósito ha sido reportado exitosamente! Espera la confirmación del administrador.")
                else:
                    messages.error(request, "Faltan datos en el formulario. Asegúrate de adjuntar el comprobante y seleccionar la cuenta destino.")
            except ValueError:
                messages.error(request, "El formato del monto ingresado es incorrecto.")
                
            return redirect('ahorro')

        # ==============================================================
        # ACCIÓN 2: SOLICITAR UN RETIRO DE FONDOS
        # ==============================================================
        elif action == 'solicitar_retiro':
            monto_retiro = request.POST.get('monto_retiro')
            try:
                monto = abs(float(monto_retiro))
                
                if 0 < monto <= float(total_ahorrado):
                    Ahorro.objects.create(
                        idsocio=socio,
                        idbingo=None,
                        tipoahorro='Voluntario',
                        montoahorro=-monto,
                        estadoahorro='Retirar', # <--- CAMBIO: Nuevo estado exclusivo para retiros
                        fechaahorro=timezone.now(),
                        origenahorro='Retiro Solicitado'
                    )
                    messages.success(request, f"Tu solicitud de retiro por ${monto:.2f} ha sido enviada a tesorería.")
                else:
                    messages.error(request, "El monto solicitado supera tu saldo disponible y acreditado.")
            except ValueError:
                messages.error(request, "Monto de retiro inválido.")
                
            return redirect('ahorro')
    contexto = {
        'socio': socio,
        'historial_ahorros': historial_ahorros,
        'total_ahorrado': total_ahorrado,
        'metodos_pago': MetodoPago.objects.filter(estadometodopago='Activo')
    }
    return render(request, 'cuentas/ahorro.html', contexto)

@login_required
def aporte_y_regalos(request):
    # 1. Validación del Socio e Identificación de su Antigüedad
    socio_obj = Socio.objects.filter(cisocio=request.user.username).first()
    fecha_registro_socio = request.user.date_joined # Obtenemos la fecha exacta en la que se creó su cuenta
    
    if not socio_obj or socio_obj.estadosocio not in ['Activo', 'Active']:
        messages.error(request, "Solo los socios activos pueden gestionar aportes y regalos.")
        return redirect('perfil')

    if request.method == 'POST':
        action = request.POST.get('action')

        # --- LÓGICA DE APORTES ---
        if action == 'registrar_aporte':
            id_bingo = request.POST.get('id_bingo')
            numero_semana = request.POST.get('numero_semana')
            monto_pagado = request.POST.get('monto_pagado')
            id_metodo_pago = request.POST.get('idmetodopago')
            comprobante = request.FILES.get('comprobanteaporte')

            if id_bingo and numero_semana and monto_pagado and id_metodo_pago and comprobante:
                bingo_obj = get_object_or_404(Bingo, pk=id_bingo)
                metodo_obj = get_object_or_404(MetodoPago, pk=id_metodo_pago) # <-- Validamos el objeto
                
                aporte_obj, creado = AporteSemanal.objects.get_or_create(
                    idsocio=socio_obj,
                    idbingo=bingo_obj,
                    numerosemana=numero_semana,
                    defaults={
                        'montoaporte': Decimal(str(monto_pagado)),
                        'idmetodopago': metodo_obj, # <-- Inyectamos la relación real
                        'comprobanteaporte': comprobante,
                        'estadoaporte': 'En Revision',
                        'fechaplanificadadada': timezone.now()
                    }
                )
                
                if not creado:
                    if aporte_obj.estadoaporte in ['Al Dia', 'En Revision']:
                        messages.warning(request, f"Ya tienes un pago reportado en estado '{aporte_obj.estadoaporte}' para la semana {numero_semana}.")
                    else:
                        # Si estaba Atrasado, se actualiza
                        aporte_obj.montoaporte = Decimal(str(monto_pagado))
                        aporte_obj.idmetodopago = metodo_obj
                        aporte_obj.comprobanteaporte = comprobante
                        aporte_obj.estadoaporte = 'En Revision'
                        aporte_obj.fechaplanificadadada = timezone.now()
                        aporte_obj.save()
                        messages.success(request, "Tu pago atrasado fue enviado a revisión exitosamente.")
                else:
                    messages.success(request, "Tu pago fue enviado a revisión exitosamente.")
            else:
                messages.error(request, "Faltan datos para procesar el aporte.")
            
            return redirect('aporte_y_regalos')

        # --- LÓGICA DE REGALOS (Se mantiene igual) ---
        elif action == 'registrar_regalo':
            nombre = request.POST.get('nombreregalo')
            descripcion = request.POST.get('descripcionregalo', '')
            valor = request.POST.get('valorregalo')
            imagen = request.FILES.get('urlimagen')
            
            try:
                Regalo.objects.create(
                    idsocio=socio_obj,
                    nombreregalo=nombre,
                    descripcionregalo=descripcion,
                    valorregalo=valor,
                    estadoregalo='Acumulado',
                    urlimagen=imagen
                )
                messages.success(request, "¡Regalo registrado exitosamente! Ha sido ingresado a la bodega virtual.")
            except Exception as e:
                messages.error(request, f"Error al registrar el regalo: {e}")
            return redirect('aporte_y_regalos')

    # 3. Preparación del contexto (GET)
    historial_aportes = AporteSemanal.objects.filter(idsocio=socio_obj).order_by('-fechaplanificadadada')
    mis_regalos = Regalo.objects.filter(idsocio=socio_obj).order_by('-fechaultimaactualizacion')
    
    # MAGIA DE FECHAS: 
    # Le damos un margen de 7 días hacia atrás por si se registró justo unos días después de iniciar el evento.
    # Excluimos "Cancelado", pero SÍ mostramos "Finalizado" por si tiene pagos pendientes de un bingo anterior.
    margen_fecha = fecha_registro_socio - timedelta(days=7)
    bingos_activos = Bingo.objects.exclude(estadobingo='Cancelado').filter(fechaprogramadabingo__gte=margen_fecha).order_by('-fechaprogramadabingo')

    context = {
        'socio': socio_obj,
        'historial_aportes': historial_aportes,
        'mis_regalos': mis_regalos,
        'bingos_activos': bingos_activos,
        'metodos_pago': MetodoPago.objects.filter(estadometodopago='Activo')
    }
    return render(request, 'cuentas/aporte_y_regalos.html', context)

@login_required
def cuenta_bancaria(request):
    """
    Vista para que el Socio registre y elimine sus cuentas bancarias.
    Solo accesible para Socios Activos, con límite de 2 cuentas.
    """
    socio = Socio.objects.filter(cisocio=request.user.username).first()
    # Verificamos si realmente es un socio activo
    es_socio = True if socio and socio.estadosocio in ['Activo', 'Active'] else False

    # Si es socio, cargamos sus cuentas; si no, lista vacía
    cuentas = CuentaBancaria.objects.filter(idsocio=socio) if socio else []

    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'agregar_cuenta':
            if not es_socio:
                messages.error(request, "Solo los socios activos pueden registrar cuentas bancarias.")
                return redirect('cuenta_bancaria')

            try:
                # Usamos los nombres EXACTOS de tu nuevo modelo
                nueva_cuenta = CuentaBancaria(
                    idsocio=socio,
                    nombrebanco=request.POST.get('nombrebanco'),
                    tipocuenta=request.POST.get('tipocuenta'),
                    numerocuenta=request.POST.get('numerocuenta'),
                    esprincipal=True if cuentas.count() == 0 else False, # La primera será la principal
                    estadocuenta='Activo'
                )
                
                # full_clean() fuerza la ejecución de tu def clean(self) en el modelo
                nueva_cuenta.full_clean() 
                nueva_cuenta.save()
                messages.success(request, "Cuenta bancaria agregada exitosamente a tu perfil.")
                
            except ValidationError as e:
                # Si se activa el bloqueo de las 2 cuentas, mostramos el error elegante
                if hasattr(e, 'message_dict'):
                    messages.error(request, "Error de validación en los datos.")
                else:
                    messages.error(request, e.messages[0])
            except Exception as e:
                messages.error(request, f"Error al registrar cuenta: {str(e)}")
            
            return redirect('cuenta_bancaria')
            
        elif action == 'eliminar_cuenta':
            id_cuenta = request.POST.get('id_cuenta')
            cuenta = CuentaBancaria.objects.filter(idcuentabancaria=id_cuenta, idsocio=socio).first()
            if cuenta:
                cuenta.delete()
                messages.success(request, "La cuenta bancaria ha sido eliminada.")
            return redirect('cuenta_bancaria')

    contexto = {
        'socio': socio,
        'es_socio': es_socio,
        'cuentas': cuentas
    }
    return render(request, 'cuentas/cuenta_bancaria.html', contexto)

@login_required
def pago(request):
    """
    Vista para que el Socio reporte o registre un pago realizado a sus préstamos.
    """
    socio = Socio.objects.filter(cisocio=request.user.username).first()
    
    # 1. Validación de seguridad (Adaptado a tu lógica): Solo socios activos
    if not socio or socio.estadosocio not in ['Activo', 'Active']:
        messages.warning(request, "Debes ser un Socio activo para registrar pagos de préstamos.")
        return redirect('inicio')

    # Préstamos que aún tienen saldo pendiente
    prestamos_activos = Prestamo.objects.filter(idsocio=socio).exclude(estadoprestamo='Liquidado')
    metodos_activos = MetodoPago.objects.filter(estadometodopago='Activo')
    historial_pagos = Pago.objects.filter(idprestamo__idsocio=socio).order_by('-fechapago')

    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'registrar_pago':
            # Nombres exactos de los campos del formulario HTML
            id_prestamo = request.POST.get('id_prestamo')
            id_metodo = request.POST.get('id_metodo_pago') 
            monto_pagado = request.POST.get('monto_pagado')
            referencia_pago = request.POST.get('numeroreferencia')
            # Capturamos la imagen del comprobante
            imagen_comprobante = request.FILES.get('imagencomprobante')

            try:
                monto = float(monto_pagado)
                # Validar que el préstamo pertenece al socio y existe
                prestamo = prestamos_activos.filter(idprestamo=id_prestamo).first()
                metodo = metodos_activos.filter(idmetodopago=id_metodo).first()

                # Validamos que exista la imagen y que los montos sean mayores a 0
                if prestamo and metodo and monto > 0:
                    if imagen_comprobante:
                        Pago.objects.create(
                            idprestamo=prestamo,
                            idmetodopago=metodo,
                            montopagado=monto, 
                            numeroreferencia=referencia_pago,
                            comprobantepago=imagen_comprobante, 
                            fechapago=timezone.now(),
                            estadopago='Pendiente' # Queda pendiente de verificación
                        )
                        messages.success(request, "Tu comprobante de pago ha sido registrado. Un administrador verificará la transacción pronto.")
                    else:
                        messages.error(request, "Es obligatorio subir una foto o captura del comprobante de pago.")
                else:
                    messages.error(request, "Datos inválidos. Por favor, verifica el préstamo y el método de pago seleccionado.")
            except (ValueError, TypeError):
                messages.error(request, "El monto ingresado no es válido.")
        
        return redirect('pago')

    contexto = {
        'socio': socio,
        'prestamos_activos': prestamos_activos,
        'metodos_activos': metodos_activos,
        'historial_pagos': historial_pagos
    }
    return render(request, 'negocio/pago.html', contexto)
# ===============================================================================================================================================

# ===============================================================================================================================================
# 4. ADMINISTRADOR (Consolas de Mando)
# ===============================================================================================================================================
@login_required
def dashboard(request):
    if not request.user.is_staff:
        messages.error(request, "Acceso exclusivo para el personal de administración.")
        return redirect('inicio')

    # ====================================================================
    try:
        # 1. Buscamos todas las rondas cuyo premio ya fue entregado al ganador
        partidas_entregadas = PartidaBingo.objects.filter(estadopremiomaterial='Entregado')
        for pt in partidas_entregadas:
            # 2. Si el regalo está enlazado pero atascado, lo forzamos a Entregado
            Regalo.objects.filter(idpartida=pt, estadoregalo__in=['Acumulado', 'Sorteado']).update(
                estadoregalo='Entregado', fechaultimaactualizacion=timezone.now()
            )
            # 3. Rescate: Si el regalo perdió su enlace (escrito manual), lo buscamos por nombre
            if pt.premiomaterial and pt.premiomaterial not in ['Ninguno', '[POZO_MAYOR]']:
                nombres = [n.strip() for n in pt.premiomaterial.split('+') if n.strip()]
                for nombre in nombres:
                    perdidos = Regalo.objects.filter(
                        nombreregalo__icontains=nombre, 
                        estadoregalo__in=['Acumulado', 'Sorteado']
                    )
                    for r in perdidos:
                        r.estadoregalo = 'Entregado'
                        r.idpartida = pt # Repara el enlace roto para que salga el texto "Ronda 1"
                        if hasattr(r, 'fechaentregaregalo'): 
                            r.fechaentregaregalo = timezone.now()
                        r.fechaultimaactualizacion = timezone.now()
                        r.save()
    except Exception as e:
        print(f"Error en auto-reparador: {e}")
    # ====================================================================

    if request.method == 'POST':
        action = request.POST.get('action')
        try:
            if action == 'crear_tiposocio':
                TipoSocio.objects.create(nombretiposocio=request.POST.get('nombretiposocio'), roltiposocio=request.POST.get('roltiposocio'), descripciondetiposocio=request.POST.get('descripciondetiposocio'))
                messages.success(request, "Tipo de Socio creado correctamente.")
            elif action == 'eliminar_tiposocio':
                TipoSocio.objects.get(idtiposocio=request.POST.get('id_tipo')).delete()
                messages.success(request, "Tipo de Socio eliminado.")

            # =======================================================
            # NUEVO: ACCIÓN PARA VALIDAR AL SOCIO DESDE EL DASHBOARD
            # =======================================================
            elif action == 'validar_socio':
                socio_validar = Socio.objects.get(idsocio=request.POST.get('id_socio'))
                socio_validar.estadosocio = 'Activo'
                socio_validar.save()
                messages.success(request, f"¡Socio {socio_validar.primernombresocio} validado y activado correctamente en la cooperativa!")
                
            # =======================================================
            # NUEVO: ACCIÓN PARA RECHAZAR AL SOCIO DESDE EL DASHBOARD
            # =======================================================
            elif action == 'rechazar_socio':
                socio_rechazar = Socio.objects.get(idsocio=request.POST.get('id_socio'))
                nombre_rechazado = socio_rechazar.primernombresocio
                
                usuario_asociado = User.objects.filter(username=socio_rechazar.cisocio).first()
                jugador_asociado = Jugador.objects.filter(idsocio=socio_rechazar).first()
                
                if jugador_asociado:
                    # Si es un jugador que pidió ascenso, le quitamos la vinculación
                    # y le devolvemos sus nombres para no destruir su cuenta de juego
                    jugador_asociado.idsocio = None
                    jugador_asociado.nombresjugador = socio_rechazar.primernombresocio
                    jugador_asociado.apellidosjugador = socio_rechazar.primerapellidosocio
                    jugador_asociado.save()
                    
                # Eliminamos la solicitud de socio
                socio_rechazar.delete()
                
                # SOLO eliminamos las credenciales de Django si no hay un jugador de por medio
                if usuario_asociado and not jugador_asociado:
                    usuario_asociado.delete()
                    
                messages.warning(request, f"La solicitud de Socio de {nombre_rechazado} ha sido rechazada.")
                
            elif action == 'recarga_masiva_credito':
                ids_str = request.POST.get('jugadores_ids', '')
                tipo_saldo = request.POST.get('tipo_saldo')
                monto_str = request.POST.get('monto', '0.00')
                
                if not ids_str:
                    messages.error(request, "Error: No seleccionaste ningún jugador para la recarga.")
                    return redirect('dashboard')
                
                # Convertimos la cadena de IDs separados por comas en una lista limpia de enteros
                ids_lista = [int(x) for x in ids_str.split(',') if x.strip().isdigit()]
                monto = Decimal(monto_str)
                
                if monto <= 0:
                    messages.error(request, "El monto a inyectar debe ser mayor a cero.")
                    return redirect('dashboard')

                
                if tipo_saldo == 'real':
                    Jugador.objects.filter(idjugador__in=ids_lista).update(
                        saldocreditojugador=F('saldocreditojugador') + monto
                    )
                    messages.success(request, f"⚡ Inyección Masiva: Se añadieron ${monto} (Saldo Real) a {len(ids_lista)} jugadores.")
                elif tipo_saldo == 'virtual':
                    Jugador.objects.filter(idjugador__in=ids_lista).update(
                        saldovirtualjugador=F('saldovirtualjugador') + monto
                    )
                    messages.success(request, f"⚡ Inyección Masiva: Se añadieron {monto} Puntos (Saldo Virtual) a {len(ids_lista)} jugadores.")
                else:
                    messages.error(request, "Parámetro de saldo no reconocido.")
            elif action == 'crear_plataforma':
                estado_plat = True if request.POST.get('estadoplataforma') == 'on' else False
                PlataformaJuego.objects.create(nombreplataforma=request.POST.get('nombreplataforma'), urlplataforma=request.POST.get('urlplataforma'), descripcionplataforma=request.POST.get('descripcionplataforma'), contactoplataforma=request.POST.get('contactoplataforma'), estadoplataforma=estado_plat, fechaadquisicionlicencia=request.POST.get('fechaadquisicionlicencia') or None, fechavencimientolicencia=request.POST.get('fechavencimientolicencia') or None, logoplataforma=request.FILES.get('logoplataforma'))
                messages.success(request, "Plataforma de Juego registrada con éxito.")
            elif action == 'eliminar_plataforma':
                PlataformaJuego.objects.get(idplataformajuego=request.POST.get('id_plataforma')).delete()
                messages.success(request, "Plataforma eliminada del sistema.")
            elif action == 'crear_bingo':
                unidad_venta = get_object_or_404(UnidadMonetaria, idunidadmonetaria=request.POST.get('idunidad_venta'))
                
                # FIX: Manejo inteligente si el Pozo Mayor viene desactivado desde el HTML
                id_premio = request.POST.get('idunidad_premio')
                if id_premio:
                    unidad_premio = get_object_or_404(UnidadMonetaria, idunidadmonetaria=id_premio)
                    premio_mayor = request.POST.get('premiomayor', 0)
                    desc_premio = request.POST.get('descripcionpremiomayor', 'Ninguno')
                else:
                    unidad_premio = unidad_venta # Usa la misma moneda de venta como relleno
                    premio_mayor = 0
                    desc_premio = 'Sin Pozo Mayor'

                Bingo.objects.create(
                    idunidad_venta=unidad_venta,
                    idunidad_premio=unidad_premio, 
                    titulobingo=request.POST.get('titulobingo'), 
                    fechaprogramadabingo=request.POST.get('fechaprogramadabingo'), 
                    tipobingo=request.POST.get('tipobingo'), 
                    lugarbingo=request.POST.get('lugarbingo'), 
                    urlsesionbingo=request.POST.get('urlsesionbingo'), 
                    preciocarton=request.POST.get('preciocarton'), 
                    premiomayor=premio_mayor, 
                    descripcionpremiomayor=desc_premio, 
                    estadobingo=request.POST.get('estadobingo'), 
                    descripcionpremios=request.POST.get('descripcionpremios', ''), 
                    rutaimagenpremiomayor=request.FILES.get('rutaimagenpremiomayor'), 
                    urlvideopromocional=request.FILES.get('urlvideopromocional')
                )
                messages.success(request, "¡Jornada de Bingo creada exitosamente!")

            elif action == 'revelar_bingo':
                bingo_id = request.POST.get('id_bingo')
                bingo = get_object_or_404(Bingo, idbingo=bingo_id)
                bingo.estadobingo = 'Programado'
                bingo.save()
                messages.success(request, f"¡El bingo '{bingo.titulobingo}' ahora está público y visible para los usuarios!")
            
            elif action == 'gestionar_aporte':
                ids_str = request.POST.get('id_aporte', '')
                ids_aportes = [int(x) for x in ids_str.split(',') if x.strip().isdigit()]
                decision = request.POST.get('decision')
                
                for id_aporte in ids_aportes:
                    aporte = get_object_or_404(AporteSemanal, pk=id_aporte)
                    if decision == 'Aprobar':
                        aporte.estadoaporte = 'Al Dia'
                        aporte.fechaplanificadadada = timezone.now()
                        aporte.save()
                    elif decision == 'Rechazar':
                        aporte.estadoaporte = 'Atrasado'
                        aporte.save()
                        
                if decision == 'Aprobar':
                    messages.success(request, f"¡{len(ids_aportes)} aportes validados exitosamente!")
                else:
                    messages.warning(request, f"{len(ids_aportes)} aportes marcados como atrasados.")

            # =======================================================
            # 1. GESTIÓN DE CRÉDITOS (APROBAR / RECHAZAR)
            # =======================================================
            elif action == 'gestionar_credito':
                prestamo = get_object_or_404(Prestamo, pk=request.POST.get('id_prestamo'))
                decision = request.POST.get('decision')
                
                if decision == 'Aprobar':
                    # Al aprobar y desembolsar, el préstamo entra en etapa de pago por cuotas
                    prestamo.estadoprestamo = 'En Curso'
                    prestamo.save()
                    messages.success(request, f"¡Crédito #{prestamo.idprestamo} aprobado y puesto 'En Curso' para {prestamo.idsocio.primernombresocio}!")
                
                elif decision == 'Rechazar':
                    prestamo.estadoprestamo = 'Rechazado'
                    prestamo.save()
                    messages.warning(request, f"El crédito #{prestamo.idprestamo} ha sido rechazado.")

            # =======================================================
            # 2. GESTIÓN DE AMORTIZACIONES Y PAGOS POR CUOTAS
            # =======================================================
            elif action == 'gestionar_transaccion':
                id_pago_str = request.POST.get('id_pago', '')
                id_transaccion_str = request.POST.get('id_transaccion', '')
                decision = request.POST.get('decision')

                if id_pago_str:
                    ids_pagos = [int(x) for x in id_pago_str.split(',') if x.strip().isdigit()]
                    for id_pago in ids_pagos:
                        pago = get_object_or_404(Pago, pk=id_pago)
                        prestamo = pago.idprestamo
                        
                        if decision == 'Aprobar' and pago.estadopago == 'Pendiente':
                            pago.estadopago = 'Validado'
                            pago.fechaconfirmacionadmin = timezone.now()
                            pago.save()
                            prestamo.saldopendiente -= pago.montopagado
                            
                            if prestamo.saldopendiente <= 0:
                                prestamo.saldopendiente = Decimal('0.00')
                                prestamo.estadoprestamo = 'Liquidado'
                            else:
                                prestamo.estadoprestamo = 'En Curso'
                            prestamo.save()
                            
                        elif decision == 'Rechazar' and pago.estadopago == 'Pendiente':
                            pago.estadopago = 'Rechazado'
                            pago.fechaconfirmacionadmin = timezone.now()
                            pago.save()
                    messages.success(request, f"¡Operación masiva procesada para {len(ids_pagos)} amortizaciones!")

                elif id_transaccion_str:
                    ids_transacciones = [int(x) for x in id_transaccion_str.split(',') if x.strip().isdigit()]
                    for id_transaccion in ids_transacciones:
                        txn = get_object_or_404(TransaccionRecarga, pk=id_transaccion)
                        if decision == 'Aprobar' and txn.estado == 'Pendiente':
                            txn.estado = 'Completada'
                            txn.fechaactualizacion = timezone.now()
                            jugador_txn = txn.idjugador
                            if txn.idtarjeta.tiposaldo == 'Efectivo':
                                jugador_txn.saldocreditojugador += txn.saldo_acreditar
                            else:
                                jugador_txn.saldovirtualjugador += txn.saldo_acreditar
                            jugador_txn.save()
                            txn.save()
                        elif decision == 'Rechazar' and txn.estado == 'Pendiente':
                            txn.estado = 'Rechazada'
                            txn.fechaactualizacion = timezone.now()
                            txn.save()
                    messages.success(request, f"¡Operación masiva procesada para {len(ids_transacciones)} recargas de billetera!")

            elif action == 'aperturar_mes_cobros':
                mes = int(request.POST.get('mes_generar'))
                anio = int(request.POST.get('anio_generar'))
                dia_corte_esperado = int(request.POST.get('dia_corte')) # 0=Lunes, 6=Domingo
                
                try:
                    # Traemos a todos los socios activos de la cooperativa
                    socios_activos = Socio.objects.filter(estadosocio='Activo')
                    aportes_a_crear = []
                    
                    # Generamos automáticamente 4 semanas de aportes
                    for semana_num in range(1, 5):
                        # Calculamos una fecha base estimada para cada semana del mes
                        dia_estimado = 1 + (semana_num - 1) * 7
                        try:
                            fecha_base = timezone.make_aware(datetime(anio, mes, dia_estimado))
                        except ValueError:
                            # Fallback si el día se pasa del fin de mes
                            fecha_base = timezone.make_aware(datetime(anio, mes, 28))
                            
                        # Ajustamos la fecha al "Día de Corte" seleccionado
                        dias_diferencia = dia_corte_esperado - fecha_base.weekday()
                        fecha_corte = fecha_base + timedelta(days=dias_diferencia)
                        
                        # Generamos los registros de la deuda (Pendiente) para todos los socios
                        for socio in socios_activos:
                            # Verificamos que no exista previamente para no duplicar deudas
                            if not AporteSemanal.objects.filter(
                                idsocio=socio, 
                                numerosemana=semana_num, 
                                fechaplanificadadada__year=anio, 
                                fechaplanificadadada__month=mes
                            ).exists():
                                
                                aportes_a_crear.append(AporteSemanal(
                                    idsocio=socio,
                                    numerosemana=semana_num,
                                    montoaporte=Decimal('0.00'), # Monto base de deuda
                                    estadoaporte='Pendiente',
                                    fechaplanificadadada=fecha_corte
                                ))
                    
                    if aportes_a_crear:
                        # bulk_create inyecta todo de golpe a la base de datos (más rápido)
                        AporteSemanal.objects.bulk_create(aportes_a_crear)
                        messages.success(request, f"¡Éxito! Se generaron las obligaciones de cobro para {mes}/{anio} a todos los socios activos.")
                    else:
                        messages.warning(request, f"Las semanas para {mes}/{anio} ya estaban generadas o no hay socios activos en el sistema.")
                        
                except Exception as e:
                    messages.error(request, f"Error al generar el mes de cobros: {str(e)}")

            elif action == 'editar_bingo':
                bingo = Bingo.objects.get(idbingo=request.POST.get('id_bingo'))
                bingo.idunidad_venta = get_object_or_404(UnidadMonetaria, idunidadmonetaria=request.POST.get('idunidad_venta'))
                
                id_premio = request.POST.get('idunidad_premio')
                if id_premio:
                    bingo.idunidad_premio = get_object_or_404(UnidadMonetaria, idunidadmonetaria=id_premio)
                    bingo.premiomayor = request.POST.get('premiomayor', 0)
                    bingo.descripcionpremiomayor = request.POST.get('descripcionpremiomayor', '')
                else:
                    bingo.idunidad_premio = bingo.idunidad_venta
                    bingo.premiomayor = 0
                    bingo.descripcionpremiomayor = 'Sin Pozo Mayor'
                    
                bingo.titulobingo = request.POST.get('titulobingo')
                bingo.preciocarton = request.POST.get('preciocarton')
                bingo.descripcionpremios = request.POST.get('descripcionpremios', '')
                
                if request.POST.get('fechaprogramadabingo'): 
                    bingo.fechaprogramadabingo = request.POST.get('fechaprogramadabingo')
                
                bingo.tipobingo = request.POST.get('tipobingo')
                bingo.lugarbingo = request.POST.get('lugarbingo')
                bingo.urlsesionbingo = request.POST.get('urlsesionbingo')
                
                estado_anterior = bingo.estadobingo
                nuevo_estado = request.POST.get('estadobingo')
                bingo.estadobingo = nuevo_estado
                
                if 'rutaimagenpremiomayor' in request.FILES: 
                    bingo.rutaimagenpremiomayor = request.FILES['rutaimagenpremiomayor']
                if 'urlvideopromocional' in request.FILES: 
                    bingo.urlvideopromocional = request.FILES['urlvideopromocional']
                
                bingo.save()
                
                # ==========================================================
                # ÁRBITRO DIGITAL: DISPARO INICIAL
                # ==========================================================
                if nuevo_estado == 'En Curso' and estado_anterior != 'En Curso':
                    primera_partida = PartidaBingo.objects.filter(idbingo=bingo).order_by('idpartidabingo').first()
                    if primera_partida and primera_partida.estadopartida == 'Programada':
                        primera_partida.estadopartida = 'En Juego'
                        primera_partida.horainiciopartida = timezone.now()
                        primera_partida.save()
                        messages.success(request, "¡Bingo iniciado! La primera ronda ha comenzado automáticamente.")
                # ==========================================================

                if nuevo_estado == 'Finalizado' and estado_anterior != 'Finalizado':
                    cartones_temporales = CartonPartidaBingo.objects.filter(idpartida__idbingo=bingo, idcarton__esmaestro=False).values_list('idcarton', flat=True)
                    ids_a_borrar = list(set(cartones_temporales))
                    if ids_a_borrar:
                        CartonPartidaBingo.objects.filter(idpartida__idbingo=bingo, idcarton__esmaestro=False).delete()
                        Carton.objects.filter(idcarton__in=ids_a_borrar).delete()
                        messages.success(request, f"¡Bingo Finalizado! El sistema ha autodestruido {len(ids_a_borrar)} cartones temporales.")
                    else:
                        messages.success(request, "Jornada de Bingo actualizada y Finalizada correctamente.")
                else:
                    messages.success(request, "Jornada de Bingo actualizada correctamente.")
                
            elif action == 'eliminar_bingo':
                Bingo.objects.get(idbingo=request.POST.get('id_bingo')).delete()
                messages.success(request, "Jornada de Bingo eliminada por completo.")
                
            elif action == 'crear_partida':
                bingo_obj = Bingo.objects.get(idbingo=request.POST.get('idbingo'))
                
                es_pozo_mayor = request.POST.get('es_pozo_mayor') == 'on'
                tipo_premio = request.POST.get('tipo_premio') 
                
                if es_pozo_mayor:
                    valor_premio = 0
                    premio_material = '[POZO_MAYOR]'
                else:
                    if tipo_premio == 'dinero':
                        valor_premio = request.POST.get('valorpremio', 0)
                        premio_material = 'Ninguno'
                    elif tipo_premio == 'fisico':
                        valor_premio = request.POST.get('valorpremio_fisico', 0)
                        premio_material = request.POST.get('premiomaterial', 'Ninguno')
                    elif tipo_premio == 'regalos':
                        regalos_ids = request.POST.getlist('regalos_ids')
                        regalos_seleccionados = Regalo.objects.filter(idregalo__in=regalos_ids)
                        
                        valor_premio = sum(r.valorregalo for r in regalos_seleccionados) if regalos_seleccionados else 0
                        nombres_regalos = [r.nombreregalo for r in regalos_seleccionados]
                        premio_material = " + ".join(nombres_regalos) if nombres_regalos else 'Regalos Sorpresa'
                    else:
                        valor_premio = request.POST.get('valorpremio', 0)
                        premio_material = request.POST.get('premiomaterial', 'Ninguno')
                        
                    if not valor_premio or str(valor_premio).strip() == '': valor_premio = 0
                    if not premio_material or str(premio_material).strip() == '': premio_material = 'Ninguno'
                
                # [EN LA SECCIÓN DE crear_partida] ...
                # SE CREA LA RONDA
                nueva_partida = PartidaBingo.objects.create(
                    idbingo=bingo_obj, 
                    nombreronda=request.POST.get('nombreronda'), 
                    modalidad_victoria=request.POST.get('modalidad_victoria', 'Tabla Llena'),
                    valorpremio=valor_premio, 
                    premiomaterial=premio_material, 
                    estadopartida='Programada', 
                    bolascantadas='', 
                    ultimabola=0 
                )
                
                # MAGIA PRIORIDAD 4: Vinculamos los regalos a la nueva ronda (USANDO .save() PARA FORZAR EL ENLACE)
                if not es_pozo_mayor and tipo_premio == 'regalos' and 'regalos_ids' in locals():
                    for r_id in regalos_ids:
                        regalo = Regalo.objects.filter(idregalo=r_id).first()
                        if regalo:
                            regalo.estadoregalo = 'Sorteado'
                            regalo.fechaultimaactualizacion = timezone.now()
                            regalo.idpartida = nueva_partida
                            regalo.save()
                
                if es_pozo_mayor:
                    messages.success(request, f"¡Ronda '{request.POST.get('nombreronda')}' aperturada! Jugarán por el POZO MAYOR de ${bingo_obj.premiomayor}.")
                else:
                    messages.success(request, f"¡Ronda '{request.POST.get('nombreronda')}' aperturada con modalidad {request.POST.get('modalidad_victoria')}!")
                
            # =======================================================
            # LOGÍSTICA DE ENTREGA DE PREMIOS FÍSICOS (CORREGIDA)
            # =======================================================
            elif action == 'entregar_premio_fisico':
                partida = PartidaBingo.objects.get(idpartidabingo=request.POST.get('id_partida'))
                partida.estadopremiomaterial = 'Entregado'
                partida.save()
                
                # 1. Buscamos primero por relación directa en la Base de Datos
                regalos_vinculados = list(Regalo.objects.filter(idpartida=partida))
                
                # 2. FALLBACK ROBUSTO: Si se rompió el enlace, buscamos por aproximación de nombre
                if not regalos_vinculados and partida.premiomaterial and partida.premiomaterial != 'Ninguno':
                    nombres_regalos = [n.strip() for n in partida.premiomaterial.split('+') if n.strip()]
                    
                    filtro_nombres = Q()
                    for nombre in nombres_regalos:
                        filtro_nombres |= Q(nombreregalo__icontains=nombre)
                        
                    # Buscamos cualquier regalo que coincida con el nombre
                    regalos_vinculados = list(Regalo.objects.filter(filtro_nombres))
                
                # 3. Actualizamos y reparamos forzosamente todos los regalos encontrados
                for r in regalos_vinculados:
                    r.estadoregalo = 'Entregado'
                    r.fechaultimaactualizacion = timezone.now()
                    if hasattr(r, 'fechaentregaregalo'):
                        r.fechaentregaregalo = timezone.now()
                    r.idpartida = partida # Repara el enlace roto para que salga el nombre de la ronda
                    r.save()
                
                messages.success(request, f"¡Excelente! El premio físico '{partida.premiomaterial}' ha sido marcado como ENTREGADO y sincronizado en bodega.")
                
            elif action == 'editar_partida':
                partida = PartidaBingo.objects.get(idpartidabingo=request.POST.get('id_partida'))
                partida.nombreronda = request.POST.get('nombreronda')
                partida.modalidad_victoria = request.POST.get('modalidad_victoria')
                
                es_pozo_mayor = request.POST.get('es_pozo_mayor') == 'on'
                tipo_premio = request.POST.get('tipo_premio') 
                
                if es_pozo_mayor:
                    partida.valorpremio = 0
                    partida.premiomaterial = '[POZO_MAYOR]'
                else:
                    if tipo_premio == 'dinero':
                        partida.valorpremio = request.POST.get('valorpremio', 0)
                        partida.premiomaterial = 'Ninguno'
                    elif tipo_premio == 'fisico':
                        partida.valorpremio = request.POST.get('valorpremio_fisico', 0)
                        partida.premiomaterial = request.POST.get('premiomaterial', 'Ninguno')
                    elif tipo_premio == 'regalos':
                        regalos_ids = request.POST.getlist('regalos_ids')
                        regalos_seleccionados = Regalo.objects.filter(idregalo__in=regalos_ids)
                        
                        partida.valorpremio = sum(r.valorregalo for r in regalos_seleccionados) if regalos_seleccionados else 0
                        nombres_regalos = [r.nombreregalo for r in regalos_seleccionados]
                        partida.premiomaterial = " + ".join(nombres_regalos) if nombres_regalos else 'Regalos Sorpresa'
                        
                        # Liberamos los regalos viejos que tenía esta ronda
                        Regalo.objects.filter(idpartida=partida).update(estadoregalo='Acumulado', idpartida=None, fechaultimaactualizacion=timezone.now())
                        # Enlazamos los nuevos
                        for r_id in regalos_ids:
                            regalo = Regalo.objects.filter(idregalo=r_id).first()
                            if regalo:
                                regalo.estadoregalo = 'Sorteado'
                                regalo.fechaultimaactualizacion = timezone.now()
                                regalo.idpartida = partida
                                regalo.save()
                    else:
                        partida.valorpremio = request.POST.get('valorpremio', 0)
                        partida.premiomaterial = request.POST.get('premiomaterial', 'Ninguno')
                        
                    if not partida.valorpremio or str(partida.valorpremio).strip() == '': partida.valorpremio = 0
                    if not partida.premiomaterial or str(partida.premiomaterial).strip() == '': partida.premiomaterial = 'Ninguno'
                
                partida.save()
                messages.success(request, f"¡Ronda '{partida.nombreronda}' actualizada correctamente!")

            elif action == 'eliminar_partida':
                PartidaBingo.objects.get(idpartidabingo=request.POST.get('id_partida')).delete()
                messages.success(request, "Ronda eliminada de forma segura.")
            # =======================================================
            elif action == 'editar_configuracion':
                config, created = ConfiguracionWeb.objects.get_or_create(idconfiguracion=1)
                config.titulosobrenosotros = request.POST.get('titulosobrenosotros', config.titulosobrenosotros)
                config.descripcionsobrenosotros = request.POST.get('descripcionsobrenosotros', config.descripcionsobrenosotros)
                config.numerowhatsapp = request.POST.get('numerowhatsapp', config.numerowhatsapp)
                config.enlaceinstagram = request.POST.get('enlaceinstagram', config.enlaceinstagram)
                config.enlacefacebook = request.POST.get('enlacefacebook', config.enlacefacebook)
                if 'imagenpromocional' in request.FILES: config.imagenpromocional = request.FILES['imagenpromocional']
                config.save()
                messages.success(request, "Configuración del sitio web actualizada correctamente.")
            elif action == 'generar_cartones':
                cantidad = int(request.POST.get('cantidad_cartones', 0))
                if cantidad > 0:
                    # 1. Generamos la matemática de los cartones
                    lote = generar_lote_cartones(cantidad)
                    
                    # 2. Preparamos los objetos para la base de datos (ESTO ERA LO QUE FALTABA)
                    cartones_db = [
                        Carton(
                            codigocarton=c['codigo'], 
                            matriznumeros=c['matriz'], 
                            esmaestro=True
                        ) for c in lote
                    ]
                    
                    # 3. Guardamos todos los cartones de golpe en la base de datos
                    Carton.objects.bulk_create(cartones_db)
                    
                    messages.success(request, f"¡Fábrica terminada! Se han estampado y guardado {cantidad} cartones exitosamente.")
            elif action == 'eliminar_carton':
                Carton.objects.get(idcarton=request.POST.get('id_carton')).delete()
                messages.success(request, "Cartón retirado del inventario general.")
            elif action == 'editar_socio':
                s = Socio.objects.get(idsocio=request.POST.get('id_socio'))
                
                # Actualizamos todos los parámetros de la base
                s.primernombresocio = request.POST.get('primer_nombre')
                s.segundonombresocio = request.POST.get('segundo_nombre', '')
                s.primerapellidosocio = request.POST.get('primer_apellido')
                s.segundoapellidosocio = request.POST.get('segundo_apellido', '')
                s.cisocio = request.POST.get('cedula')
                s.fechanacimientosocio = request.POST.get('fecha_nacimiento')
                s.nacionalidad = request.POST.get('nacionalidad')
                s.sexosocio = request.POST.get('sexo')
                s.telefonopersonalsocio = request.POST.get('telefono')
                s.telefonotrabajosocio = request.POST.get('telefonofijo', '')
                s.correosocio = request.POST.get('correo', '')
                s.direcciondomiciliosocio = request.POST.get('direccion')
                s.direcciontrabajosocio = request.POST.get('direcciontrabajo', '')
                s.estadosocio = request.POST.get('estado')
                
                tipo = request.POST.get('id_tipo_socio')
                if tipo: s.idtiposocio_id = tipo
                s.save()
                
                # Sincronizamos las credenciales del usuario en Django
                user_s = User.objects.filter(username=s.cisocio).first()
                pwd = request.POST.get('password_nueva')
                if user_s:
                    user_s.email = s.correosocio
                    user_s.first_name = s.primernombresocio
                    user_s.last_name = s.primerapellidosocio
                    if pwd: user_s.set_password(pwd)
                    user_s.save()
                    
                messages.success(request, f"Perfil completo de {s.primernombresocio} actualizado correctamente.")
            elif action == 'editar_jugador':
                j = Jugador.objects.get(idjugador=request.POST.get('id_jugador'))
                
                # Si es un jugador externo (no es socio), actualizamos su info personal
                if not j.idsocio:
                    j.nombresjugador = request.POST.get('nombres')
                    j.apellidosjugador = request.POST.get('apellidos')
                    j.cedulaidentidadjugador = request.POST.get('cedula')
                    j.nacionalidad = request.POST.get('nacionalidad')
                    
                # Info compartida
                j.aliasjugador = request.POST.get('alias')
                j.correojugador = request.POST.get('correo')
                j.saldocreditojugador = request.POST.get('saldo_real', j.saldocreditojugador)
                j.saldovirtualjugador = request.POST.get('saldo_virtual', j.saldovirtualjugador)
                j.estadocuentajugador = request.POST.get('estado')
                j.save()
                
                # Sincronizamos credenciales del usuario de juego
                user_j = User.objects.filter(username=j.cedulaidentidadjugador).first()
                pwd = request.POST.get('password_nueva')
                if user_j:
                    user_j.email = j.correojugador
                    if not j.idsocio: # Solo si no es socio le pisamos los nombres al Auth
                        user_j.first_name = j.nombresjugador
                        user_j.last_name = j.apellidosjugador
                    if pwd: user_j.set_password(pwd)
                    user_j.save()
                    
                messages.success(request, f"Perfil de juego y billetera de '{j.aliasjugador}' actualizados correctamente.")
            elif action == 'crear_moneda':
                estado = True if request.POST.get('estadomoneda') == 'on' else False
                UnidadMonetaria.objects.create(
                    nombremoneda=request.POST.get('nombremoneda'),
                    tipomoneda=request.POST.get('tipomoneda'),
                    simbolomoneda=request.POST.get('simbolomoneda'),
                    tasaconversionmoneda=request.POST.get('tasaconversionmoneda'),
                    estadomoneda=estado
                )
                messages.success(request, "Nueva unidad monetaria registrada con éxito.")
                
            elif action == 'editar_moneda':
                moneda = UnidadMonetaria.objects.get(idunidadmonetaria=request.POST.get('id_moneda'))
                moneda.nombremoneda = request.POST.get('nombremoneda')
                moneda.tipomoneda = request.POST.get('tipomoneda')
                moneda.simbolomoneda = request.POST.get('simbolomoneda')
                moneda.tasaconversionmoneda = request.POST.get('tasaconversionmoneda')
                moneda.estadomoneda = True if request.POST.get('estadomoneda') == 'on' else False
                moneda.save()
                messages.success(request, "Divisa actualizada correctamente.")
                
            elif action == 'eliminar_moneda':
                UnidadMonetaria.objects.get(idunidadmonetaria=request.POST.get('id_moneda')).delete()
                messages.success(request, "Divisa eliminada del sistema.")

            # =======================================================
            # NUEVO: GESTIÓN DE TIENDA Y TRANSACCIONES
            # =======================================================
            elif action == 'crear_tarjeta_recarga':
                es_pop = True if request.POST.get('espopular') == 'on' else False
                TarjetaRecarga.objects.create(
                    nombretarjetarecarga=request.POST.get('nombre'),
                    tiposaldo=request.POST.get('tiposaldo'),
                    montotarjetarecarga=request.POST.get('monto'),
                    preciodetarjetarecarga=request.POST.get('precio'),
                    descripciontarjetarecarga=request.POST.get('descripcion'),
                    estado=request.POST.get('estado'),
                    espopular=es_pop
                )
                messages.success(request, "Tarjeta de recarga añadida al catálogo de la tienda.")
                
            elif action == 'eliminar_tarjeta_recarga':
                TarjetaRecarga.objects.get(idtarjetarecarga=request.POST.get('id_tarjeta')).delete()
                messages.success(request, "Tarjeta de recarga eliminada del sistema.")

            elif action == 'gestionar_ahorro':
                ids_str = request.POST.get('id_ahorro', '')
                ids_ahorros = [int(x) for x in ids_str.split(',') if x.strip().isdigit()]
                decision = request.POST.get('decision')
                
                for id_ahorro in ids_ahorros:
                    ahorro = get_object_or_404(Ahorro, pk=id_ahorro)
                    if decision == 'Aprobar':
                        ahorro.estadoahorro = 'Acreditado'
                        ahorro.fechaultimaactualizacion = timezone.now()
                        ahorro.save()
                    elif decision == 'Rechazar':
                        ahorro.estadoahorro = 'Rechazado'
                        ahorro.save()
                        
                messages.success(request, f"¡Operación masiva procesada para {len(ids_ahorros)} transacciones de ahorro!")

            elif action == 'procesar_retiro':
                id_ahorro = request.POST.get('id_ahorro')
                decision = request.POST.get('decision')
                ahorro = get_object_or_404(Ahorro, pk=id_ahorro)
                
                if decision == 'Aprobar':
                    comprobante = request.FILES.get('comprobante_retiro')
                    ahorro.estadoahorro = 'Acreditado'
                    ahorro.fechaultimaactualizacion = timezone.now()
                    if comprobante:
                        ahorro.comprobanteahorro = comprobante
                    ahorro.save()
                    messages.success(request, f"¡Retiro de {ahorro.idsocio.primernombresocio} procesado y comprobante guardado!")
                    
                elif decision == 'Rechazar':
                    ahorro.estadoahorro = 'Rechazado'
                    ahorro.save()
                    messages.warning(request, "La solicitud de retiro ha sido denegada.")
            elif action == 'procesar_retiro':
                id_ahorro = request.POST.get('id_ahorro')
                decision = request.POST.get('decision')
                ahorro = get_object_or_404(Ahorro, pk=id_ahorro)
                
                if decision == 'Aprobar':
                    comprobante = request.FILES.get('comprobante_retiro')
                    ahorro.estadoahorro = 'Acreditado'
                    ahorro.fechaultimaactualizacion = timezone.now()
                    if comprobante:
                        ahorro.comprobanteahorro = comprobante
                    ahorro.save()
                    messages.success(request, f"¡Retiro de {ahorro.idsocio.primernombresocio} procesado y comprobante guardado!")
                    
                elif decision == 'Rechazar':
                    ahorro.estadoahorro = 'Rechazado'
                    ahorro.save()
                    messages.warning(request, "La solicitud de retiro ha sido denegada.")
                        
                messages.success(request, f"¡Operación masiva procesada para {len(ids_ahorros)} transacciones de ahorro!")
        except ProtectedError:
            messages.error(request, "⚠️ ERROR: No puedes eliminar este registro porque hay usuarios o datos vinculados a él.")
        except Exception as e:
            messages.error(request, f"Error en la operación: {str(e)}")
        return redirect('dashboard')
    # ====================================================================
    # INSERCIÓN SEGURA: MOTOR ESTADÍSTICO PARA EL DASHBOARD
    # ====================================================================
    hoy = timezone.now()
    ayer = hoy - timedelta(days=1)
    inicio_semana = hoy - timedelta(days=hoy.weekday())
    inicio_mes = hoy.replace(day=1)
    inicio_anio = hoy.replace(month=1, day=1)
    
    datos_graficos = {
        'hoy': {'socios': 0, 'jugadores': 0, 'ganancias': 0},
        'ayer': {'socios': 0, 'jugadores': 0, 'ganancias': 0},
        'semana': {'socios': 0, 'jugadores': 0, 'ganancias': 0},
        'mes': {'socios': 0, 'jugadores': 0, 'ganancias': 0},
        'anio': {'socios': 0, 'jugadores': 0, 'ganancias': 0},
    }

    try:
        # 1. Procesar Ganancias Reales (¡Usando la tabla Carton correcta!)
        cartones_db = Carton.objects.all()
        for c in cartones_db:
            # Buscamos el atributo correcto sin importar cómo se llame exactamente
            fecha_obj = getattr(c, 'fechacompra', getattr(c, 'fecha_creacion', getattr(c, 'fecha', None)))
            if fecha_obj:
                fecha = fecha_obj.date() if hasattr(fecha_obj, 'date') else fecha_obj
                monto = float(getattr(c, 'preciopagado', 0) or 0)
                
                if fecha == hoy.date(): datos_graficos['hoy']['ganancias'] += monto
                if fecha == ayer.date(): datos_graficos['ayer']['ganancias'] += monto
                if fecha >= inicio_semana.date(): datos_graficos['semana']['ganancias'] += monto
                if fecha >= inicio_mes.date(): datos_graficos['mes']['ganancias'] += monto
                if fecha >= inicio_anio.date(): datos_graficos['anio']['ganancias'] += monto

        # 2. Procesar Socios Registrados
        socios_db = Socio.objects.select_related('idusuario').all()
        for s in socios_db:
            fecha = None
            if hasattr(s, 'idusuario') and s.idusuario and hasattr(s.idusuario, 'date_joined'):
                fecha = s.idusuario.date_joined.date()
                
            if fecha:
                if fecha == hoy.date(): datos_graficos['hoy']['socios'] += 1
                if fecha == ayer.date(): datos_graficos['ayer']['socios'] += 1
                if fecha >= inicio_semana.date(): datos_graficos['semana']['socios'] += 1
                if fecha >= inicio_mes.date(): datos_graficos['mes']['socios'] += 1
                if fecha >= inicio_anio.date(): datos_graficos['anio']['socios'] += 1
            else:
                for k in datos_graficos: datos_graficos[k]['socios'] += 1

        # 3. Procesar Jugadores
        jugadores_db = Jugador.objects.all()
        for j in jugadores_db:
            fecha = None
            if hasattr(j, 'idusuario') and j.idusuario and hasattr(j.idusuario, 'date_joined'):
                fecha = j.idusuario.date_joined.date()
                
            if fecha:
                if fecha == hoy.date(): datos_graficos['hoy']['jugadores'] += 1
                if fecha == ayer.date(): datos_graficos['ayer']['jugadores'] += 1
                if fecha >= inicio_semana.date(): datos_graficos['semana']['jugadores'] += 1
                if fecha >= inicio_mes.date(): datos_graficos['mes']['jugadores'] += 1
                if fecha >= inicio_anio.date(): datos_graficos['anio']['jugadores'] += 1
            else:
                for k in datos_graficos: datos_graficos[k]['jugadores'] += 1
                
    except Exception as e:
        # Si algo explota silenciosamente, lo ignoramos para NO TIRAR LA PÁGINA
        print(f"Error en el motor del gráfico: {e}")
    # ====================================================================

    bingos_lista = list(Bingo.objects.all().order_by('-fechaprogramadabingo'))
    for b in bingos_lista:
        # =========================================================
        # CÁLCULO DEL POZO DINÁMICO (SOPORTE MULTIDIVISA - 45%)
        # =========================================================
        vendidos = CartonPartidaBingo.objects.filter(idpartida__idbingo=b).values('idcarton').distinct().count()
        
        # 1. Extraemos las tasas de conversión
        tasa_venta = float(b.idunidad_venta.tasaconversionmoneda)
        tasa_premio = float(b.idunidad_premio.tasaconversionmoneda)
        
        # 2. Convertimos TODO a Dólares
        ingreso_en_dolares = float(vendidos * b.preciocarton) * tasa_venta
        
        # 3. El 45% va directo al Pozo Mayor
        fondo_pozo_dolares = ingreso_en_dolares * 0.45
        premio_base_dolares = float(b.premiomayor) * tasa_premio
        
        # 4. Comparamos en Dólares y lo transformamos de vuelta
        if fondo_pozo_dolares > premio_base_dolares:
            b.pozo_dinamico_actual = fondo_pozo_dolares / tasa_premio
        else:
            b.pozo_dinamico_actual = float(b.premiomayor)
        # =========================================================

    # NUEVO: Calculamos lo mismo para las partidas de la tabla inferior
    partidas_lista = list(PartidaBingo.objects.select_related('idbingo').all())
    for pt in partidas_lista:
        vendidos = CartonPartidaBingo.objects.filter(idpartida__idbingo=pt.idbingo).values('idcarton').distinct().count()
        fondo = float(vendidos * pt.idbingo.preciocarton) * 0.75
        pt.idbingo.pozo_dinamico_actual = fondo if fondo > float(pt.idbingo.premiomayor) else float(pt.idbingo.premiomayor)

    anios_disponibles = AporteSemanal.objects.dates('fechaplanificadadada', 'year')
    anios_lista = sorted([d.year for d in anios_disponibles])
    if not anios_lista: 
        anios_lista = [timezone.now().year]

    contexto = {
        'total_socios': Socio.objects.count(), 'total_jugadores': Jugador.objects.count(), 'deuda_calle': Prestamo.objects.exclude(estadoprestamo='Liquidado').aggregate(total=Sum('saldopendiente'))['total'] or 0.00,
        'anios_aportes': anios_lista,
        'bingos_activos': Bingo.objects.exclude(estadobingo__in=['Finalizado', 'Cancelado']).count(), 'tipos_socio': TipoSocio.objects.all(),
        'socios': Socio.objects.all().order_by('-idsocio')[:50], 'accounts': CuentaBancaria.objects.all().select_related('idsocio'),
        'jugadores': Jugador.objects.all().order_by('-idjugador')[:50], 'prestamos': Prestamo.objects.all().order_by('-fechasolicitud')[:30],
        'pagos': Pago.objects.all().order_by('-fechapago'),
        'pagos_pendientes': Pago.objects.filter(estadopago='Pendiente').order_by('-fechapago'),
        'ahorros': Ahorro.objects.all().order_by('-fechaahorro')[:30], 'aportes_semanales': AporteSemanal.objects.all().order_by('-fechaplanificadadada')[:30],
        'bingos': Bingo.objects.all().order_by('-fechaprogramadabingo'), 'partidas': PartidaBingo.objects.all(),
        'regalos': Regalo.objects.all(), 'cartones': Carton.objects.all().order_by('-idcarton')[:50],
        'cartones_en_juego': CartonPartidaBingo.objects.all()[:50], 'plataformas': PlataformaJuego.objects.all(),
        'sesiones_monitoreo': SesionJuego.objects.all().order_by('-fechainiciosesion')[:30], 'config_web': ConfiguracionWeb.objects.first(),
        'unidades_monetarias': UnidadMonetaria.objects.filter(estadomoneda=True),
        'todas_monedas': UnidadMonetaria.objects.all(),
        'bingos': bingos_lista,
        'partidas': PartidaBingo.objects.all(),
        'partidas': partidas_lista,
        'tarjetas_recarga': TarjetaRecarga.objects.all().order_by('-estado', 'tiposaldo', 'preciodetarjetarecarga'),
        'transacciones_recarga': TransaccionRecarga.objects.all().select_related('idjugador', 'idtarjeta').order_by('-fechatransaccion')[:50],
    }
    # Agrega esto al final de tus variables de contexto
    contexto['bingos_con_pozo'] = list(PartidaBingo.objects.filter(premiomaterial='[POZO_MAYOR]').values_list('idbingo_id', flat=True))

    # NUEVO: Le mandamos la información empaquetada en JSON al gráfico del HTML
    contexto['datos_graficos_json'] = json.dumps(datos_graficos)
    
    return render(request, 'administrador/dashboard.html', contexto)

@login_required
def reporte_socios_puntuales(request):
    if not request.user.is_staff: return redirect('inicio')
    
    from openpyxl.styles import Font, PatternFill, Alignment
    import openpyxl.utils

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Socios Estrella"
    
    ws.append(['Cédula', 'Socio', 'Teléfono', 'Tipo de Socio', 'Historial de Aportes', 'Calificación'])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="312E81", fill_type="solid") # Azul corporativo
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    socios = Socio.objects.filter(estadosocio='Activo').select_related('idtiposocio')
    for s in socios:
        aportes = AporteSemanal.objects.filter(idsocio=s)
        total_aportes = aportes.count()
        aportes_al_dia = aportes.filter(estadoaporte='Al Dia').count()
        
        clasificacion = "Sin Historial"
        if total_aportes > 0:
            porcentaje = (aportes_al_dia / total_aportes) * 100
            if porcentaje == 100: clasificacion = "🌟 EXCELENTE (Aplica Descuento)"
            elif porcentaje >= 80: clasificacion = "👍 BUENO (Cumplido)"
            elif porcentaje >= 50: clasificacion = "⚠️ REGULAR (Alerta)"
            else: clasificacion = "❌ MOROSO (Riesgo Alto)"

        ws.append([
            s.cisocio,
            f"{s.primernombresocio} {s.primerapellidosocio}",
            s.telefonopersonalsocio,
            s.idtiposocio.nombretiposocio if s.idtiposocio else "No Definido",
            f"{aportes_al_dia} de {total_aportes} Al Día",
            clasificacion
        ])
        
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = max(max_len + 3, 12)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Socios_Estrella_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response

@login_required
def reporte_liquidacion_bingo(request, id_bingo):
    if not request.user.is_staff: return redirect('inicio')
    
    bingo = get_object_or_404(Bingo, idbingo=id_bingo)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Liquidación de Bingo"

    # =========================================================
    # 1. MOTOR DE DIVISAS (Alineación de Monedas)
    # =========================================================
    tasa_venta = float(bingo.idunidad_venta.tasaconversionmoneda)
    tasa_premio = float(bingo.idunidad_premio.tasaconversionmoneda)
    simbolo_venta = bingo.idunidad_venta.simbolomoneda # Extraemos el símbolo real ($ o 💎)

    ws.append(['Concepto Financiero', 'Detalle Operativo', f'Monto Total ({simbolo_venta})'])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1E1B4B", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # =========================================================
    # 2. INGRESOS BRUTOS
    # =========================================================
    cartones_vendidos = CartonPartidaBingo.objects.filter(idpartida__idbingo=bingo).values('idcarton').distinct().count()
    ingresos_totales = float(cartones_vendidos * bingo.preciocarton) # Esto está en la moneda de venta
    ingreso_dolares = ingresos_totales * tasa_venta
    
    # =========================================================
    # 3. LÓGICA DEL POZO MAYOR (El 45% vs Premio Base)
    # =========================================================
    fondo_pozo_45_dolares = ingreso_dolares * 0.45
    premio_base_dolares = float(bingo.premiomayor) * tasa_premio
    
    # ¿Qué pagamos? Comparamos en dólares para ser justos
    pozo_entregado_dolares = fondo_pozo_45_dolares if fondo_pozo_45_dolares > premio_base_dolares else premio_base_dolares
    
    # Lo convertimos a la divisa de venta para que cuadre en el Excel
    pozo_entregado_moneda_venta = pozo_entregado_dolares / tasa_venta 
        
    # =========================================================
    # 4. RONDAS MENORES Y LA MAGIA DEL DESVÍO DEL 30%
    # =========================================================
    fondo_30_teorico = ingresos_totales * 0.30 # Presupuesto en moneda de venta
    
    premios_efectivo_crudos = PartidaBingo.objects.filter(
        idbingo=bingo
    ).exclude(premiomaterial='[POZO_MAYOR]').aggregate(total=Sum('valorpremio'))['total'] or 0
    
    # CONVERSIÓN CRÍTICA: Pasamos los premios menores a la divisa de venta
    premios_menores_dolares = float(premios_efectivo_crudos) * tasa_premio
    premios_menores_moneda_venta = premios_menores_dolares / tasa_venta
    
    # Calculamos el ahorro (desvío) usando la misma divisa
    sobrante_desviado = fondo_30_teorico - premios_menores_moneda_venta
    if sobrante_desviado < 0: sobrante_desviado = 0 
    
    # =========================================================
    # 5. GANANCIA NETA FINAL (Matemáticamente balanceada)
    # =========================================================
    utilidad_neta = ingresos_totales - pozo_entregado_moneda_venta - premios_menores_moneda_venta

    # ================= ESCRITURA EN EXCEL =================
    ws.append(['INGRESOS BRUTOS', f'Recaudación por Cartones ({cartones_vendidos} cartones x {simbolo_venta}{bingo.preciocarton})', ingresos_totales])
    ws.append(['EGRESO: POZO MAYOR', 'Pago del Premio Principal (Convertido a divisa base)', -pozo_entregado_moneda_venta])
    ws.append(['EGRESO: RONDAS MENORES', 'Suma de premios menores pagados (Convertidos a divisa base)', -premios_menores_moneda_venta])
    ws.append(['', '', '']) 
    
    ws.append(['ANÁLISIS DE AHORRO (30%)', 'Presupuesto teórico destinado para rondas (30%)', fondo_30_teorico])
    ws.append(['DESVÍO A LA CASA', 'Dinero ahorrado (sobrante) por usar regalos físicos', sobrante_desviado])
    ws.append(['', '', '']) 
    
    ws.append(['UTILIDAD LÍQUIDA', 'Ganancia Neta Final de la Cooperativa (+ Desvío del 30%)', utilidad_neta])
    
    # ================= ESTILOS DE CELDAS =================
    ws[3][2].font = Font(color="FF0000") 
    ws[4][2].font = Font(color="FF0000") 
    ws[6][2].font = Font(color="0000FF") 
    ws[7][2].font = Font(bold=True, color="008000") 
    
    ws[9][0].font = Font(bold=True)
    ws[9][1].font = Font(bold=True)
    ws[9][2].font = Font(bold=True, color="008000" if utilidad_neta >= 0 else "FF0000")
    
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 65
    ws.column_dimensions['C'].width = 20

    # Formato de moneda dinámico usando el símbolo de la base de datos
    for row in range(2, 10):
        if ws.cell(row=row, column=3).value != '':
            ws.cell(row=row, column=3).number_format = f'"{simbolo_venta}"#,##0.00'

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Liquidacion_{bingo.titulobingo[:10]}_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response

@login_required
def control_aportes(request):
    # Barrera de seguridad: Solo administradores
    if not request.user.is_staff:
        messages.error(request, "Acceso exclusivo para el personal de administración.")
        return redirect('inicio')

    # ==========================================
    # 1. LÓGICA POST: REGISTRAR O ACTUALIZAR PAGO MANUALMENTE
    # ==========================================
    if request.method == 'POST':
        id_socio = request.POST.get('id_socio')
        id_bingo = request.POST.get('id_bingo')
        numero_semana = request.POST.get('numero_semana')
        monto = request.POST.get('monto')
        
        try:
            socio = get_object_or_404(Socio, idsocio=id_socio)
            bingo = get_object_or_404(Bingo, idbingo=id_bingo)
            
            # Buscar si el aporte ya existe o crearlo
            aporte, creado = AporteSemanal.objects.get_or_create(
                idsocio=socio,
                idbingo=bingo,
                numerosemana=numero_semana,
                defaults={
                    'montoaporte': Decimal(str(monto)), 
                    'estadoaporte': 'Al Dia',
                    'fechaplanificadadada': timezone.now()
                }
            )
            
            # Si existía (estaba pendiente o atrasado), se actualiza
            if not creado:
                aporte.montoaporte = Decimal(str(monto))
                aporte.estadoaporte = 'Al Dia'
                aporte.fechaplanificadadada = timezone.now()
                aporte.save()

            messages.success(request, f"Aporte de la semana {numero_semana} registrado exitosamente para {socio.primernombresocio}.")
            return redirect(f"/control_aportes/?bingo_id={id_bingo}")
            
        except Exception as e:
            messages.error(request, f"Error al registrar el aporte: {str(e)}")
            return redirect('control_aportes')

    # ==========================================
    # 2. LÓGICA GET: RENDERIZAR LA MATRIZ FINANCIERA
    # ==========================================
    id_bingo = request.GET.get('bingo_id')
    if id_bingo:
        bingo_seleccionado = Bingo.objects.filter(idbingo=id_bingo).first()
    else:
        bingo_seleccionado = Bingo.objects.filter(estadobingo='En Curso').first() or Bingo.objects.order_by('-fechaprogramadabingo').first()

    if not bingo_seleccionado:
        return render(request, 'administrador/control_aportes.html', {'error': 'No hay eventos de bingo creados.'})

    # Consultas optimizadas
    socios = Socio.objects.filter(estadosocio='Activo').order_by('primerapellidosocio', 'primernombresocio')
    aportes = AporteSemanal.objects.filter(idbingo=bingo_seleccionado).select_related('idsocio')

    # Determinar rango de semanas
    semanas_query = aportes.values_list('numerosemana', flat=True).distinct().order_by('numerosemana')
    semanas = list(semanas_query) if semanas_query.exists() else list(range(1, 6))

    # Creación de la matriz
    matriz_socios = {}
    for socio in socios:
        matriz_socios[socio.idsocio] = {
            'objeto_socio': socio,
            'semanas_data': {sem: None for sem in semanas},
            'total_acumulado': Decimal('0.00'),
            'tiene_atrasos': False
        }

    # Llenado de datos reales
    for aporte in aportes:
        s_id = aporte.idsocio_id
        if s_id in matriz_socios:
            if aporte.numerosemana in matriz_socios[s_id]['semanas_data']:
                matriz_socios[s_id]['semanas_data'][aporte.numerosemana] = {
                    'monto': aporte.montoaporte,
                    'estado': aporte.estadoaporte,
                    'id_aporte': getattr(aporte, 'idaporte', getattr(aporte, 'id', None)) # Respaldo por si varía el nombre del ID
                }
                
                if aporte.estadoaporte == 'Al Dia':
                    matriz_socios[s_id]['total_acumulado'] += Decimal(str(aporte.montoaporte))
                elif aporte.estadoaporte == 'Atrasado':
                    matriz_socios[s_id]['tiene_atrasos'] = True

    context = {
        'bingo_seleccionado': bingo_seleccionado,
        'todos_los_bingos': Bingo.objects.all().order_by('-fechaprogramadabingo'),
        'semanas': semanas,
        'matriz_socios': matriz_socios.values(), 
    }
    
    return render(request, 'administrador/control_aportes.html', context)

@login_required
def reporte_cartera_prestamos(request):
    if not request.user.is_staff: return redirect('inicio')
    
    from openpyxl.styles import Font, PatternFill, Alignment
    import openpyxl.utils

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cartera de Créditos"
    
    ws.append(['Cédula', 'Socio', 'Monto Solicitado', 'Total a Pagar', 'Saldo Pendiente', 'Estado'])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="B91C1C", fill_type="solid") # Rojo analítico financiero
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    prestamos = Prestamo.objects.all().select_related('idsocio')
    for p in prestamos:
        ws.append([
            p.idsocio.cisocio if p.idsocio else "N/A",
            f"{p.idsocio.primernombresocio} {p.idsocio.primerapellidosocio}" if p.idsocio else "Externo",
            float(p.montoprestamo or 0),
            float(p.montototalapagar or 0),
            float(p.saldovivoprestamo or 0),
            p.estadoprestamo
        ])
        
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = max(max_len + 3, 12)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Cartera_Creditos_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response

@login_required
def reporte_caja_semanal_pdf(request):
    import io
    from django.http import HttpResponse
    from django.utils import timezone
    from decimal import Decimal
    from django.db.models import Sum
    
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
    except ImportError:
        messages.error(request, "⚠️ Falta la librería de PDFs. Instálala ejecutando en tu terminal: pip install reportlab")
        return redirect('dashboard')

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=30)
    elements = []
    styles = getSampleStyleSheet()
    
    elements.append(Paragraph("Reporte de Caja Semanal - CoopBingo", styles['Title']))
    elements.append(Spacer(1, 15))

    anio_actual = timezone.now().year
    aportes_anio = AporteSemanal.objects.filter(fechaplanificadadada__year=anio_actual)
    
    pagados = aportes_anio.filter(estadoaporte__in=['Al Dia', 'Pagado'])
    pendientes = aportes_anio.filter(estadoaporte__in=['Pendiente', 'Atrasado', 'En Revision'])
    
    total_recaudado = pagados.aggregate(Sum('montoaporte'))['montoaporte__sum'] or Decimal('0.00')
    cantidad_pagados = pagados.count()
    
    total_bingo = Decimal(str(cantidad_pagados * 2.00))
    total_ahorro = total_recaudado - total_bingo
    
    elements.append(Paragraph(f"1. Resumen Financiero General ({anio_actual})", styles['Heading2']))
    
    data_resumen = [
        ['Concepto de Ingreso', 'Monto Calculado ($)'],
        ['Total Bruto Recaudado', f"${total_recaudado}"],
        ['Dinero Destinado al Pozo de Bingo ($2 x Cuota)', f"${total_bingo}"],
        ['Dinero Destinado a Bóveda de Ahorros', f"${total_ahorro}"],
    ]
    
    t_resumen = Table(data_resumen, colWidths=[350, 150])
    t_resumen.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#0d6efd")), # Encabezado Azul
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0,0), (-1,0), 12),
        ('BACKGROUND', (0,1), (-1,-1), colors.HexColor("#f8f9fa")),
        ('GRID', (0,0), (-1,-1), 1, colors.black),
    ]))
    
    elements.append(t_resumen)
    elements.append(Spacer(1, 25))
    
    elements.append(Paragraph("2. Socios con Semanas Pendientes o en Revisión", styles['Heading2']))
    
    data_pendientes = [['Nombre del Socio', 'Semana', 'Estado Actual']]
    for p in pendientes.order_by('fechaplanificadadada')[:30]:
        nombre = f"{p.idsocio.primernombresocio} {p.idsocio.primerapellidosocio}"
        data_pendientes.append([nombre, f"Semana {p.numerosemana}", p.estadoaporte])
        
    if len(data_pendientes) > 1:
        t_pendientes = Table(data_pendientes, colWidths=[250, 100, 150])
        t_pendientes.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#dc3545")), # Encabezado Rojo
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('GRID', (0,0), (-1,-1), 1, colors.black),
            ('TEXTCOLOR', (2,1), (2,-1), colors.red), # Las palabras "Pendiente" en rojo
        ]))
        elements.append(t_pendientes)
    else:
        elements.append(Paragraph("Excelente trabajo. No hay socios morosos registrados actualmente.", styles['Normal']))

    doc.build(elements)
    buffer.seek(0)
    
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="Cierre_Caja_CoopBingo_{anio_actual}.pdf"'
    return response

@login_required
def reporte_liquidacion_excel(request):
    if not request.user.is_staff: return redirect('inicio')
    
    from openpyxl.styles import Font, PatternFill, Alignment
    import openpyxl.utils

    # 1. Crear el lienzo de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Liquidación Anual"
    
    # 2. Títulos de las columnas con diseño
    ws.append(['Cédula', 'Socio', 'Total Ahorrado', 'Porcentaje Propiedad', 'Intereses Ganados', 'TOTAL A RECIBIR'])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="0D9488", fill_type="solid") # Un verde/teal elegante
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # ==============================================================
    # PASO 1: Calcular los POZOS TOTALES de la Cooperativa
    # ==============================================================
    # ADAPTACIÓN: Inyectamos el filtro de tu compañero (estadoahorro='Acreditado')
    total_ahorros_coop = Ahorro.objects.filter(estadoahorro='Acreditado').aggregate(Sum('montoahorro'))['montoahorro__sum'] or Decimal('0.00')
    
    prestamos_validos = Prestamo.objects.exclude(estadoprestamo__in=['Solicitado', 'Rechazado', 'Cancelado'])
    total_intereses_coop = Decimal(str(sum((p.montototalpagar - p.montoprestamosolicitado) for p in prestamos_validos)))
        
    # ==============================================================
    # PASO 2 y 3: Calcular el pedazo del pastel para CADA SOCIO
    # ==============================================================
    socios = Socio.objects.filter(estadosocio='Activo')
    
    for socio in socios:
        # ADAPTACIÓN: Filtramos también los ahorros individuales por 'Acreditado'
        ahorro_socio = Ahorro.objects.filter(idsocio=socio, estadoahorro='Acreditado').aggregate(Sum('montoahorro'))['montoahorro__sum'] or Decimal('0.00')
        
        porcentaje = Decimal('0.00')
        if total_ahorros_coop > 0:
            porcentaje = (ahorro_socio / total_ahorros_coop) * Decimal('100.00')
            
        ganancia = (porcentaje / Decimal('100.00')) * total_intereses_coop
        total_recibir = ahorro_socio + ganancia
        
        # Insertamos como float() para que Excel los reconozca como números y se puedan sumar
        ws.append([
            socio.cisocio,
            f"{socio.primerapellidosocio} {socio.primernombresocio}",
            float(ahorro_socio),
            float(porcentaje),
            float(ganancia),
            float(total_recibir)
        ])
        
    # 4. Fila final de COMPROBACIÓN
    ws.append([]) # Fila vacía para separar
    ws.append([
        '', 
        'SUMA TOTAL DE LA COOPERATIVA', 
        float(total_ahorros_coop), 
        100.00, 
        float(total_intereses_coop), 
        float(total_ahorros_coop + total_intereses_coop)
    ])
    
    # 5. Diseño: Aplicar negrita a la última fila de totales
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
        if isinstance(cell.value, float):
            cell.font = Font(bold=True, color="008000") # Totales en verde

    # 6. Diseño: Formato de moneda nativo de Excel y Porcentaje
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=3).number_format = '"$"#,##0.00'
        ws.cell(row=row, column=4).number_format = '0.00"%"'
        ws.cell(row=row, column=5).number_format = '"$"#,##0.00'
        ws.cell(row=row, column=6).number_format = '"$"#,##0.00'

    # 7. Diseño: Auto-ajuste de columnas
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = max(max_len + 2, 15)

    # Devolver el archivo como .xlsx real
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Liquidacion_Fin_Anio_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    
    wb.save(response)
    return response

@login_required
def metodos_pago(request):
    """
    Endpoint de procesamiento para los métodos de pago.
    Recibe las peticiones del dashboard, ejecuta la acción en la BD y redirige.
    """
    # Barrera de seguridad: Solo administradores pueden gestionar esto
    if not request.user.is_staff:
        messages.error(request, "Acceso denegado.")
        return redirect('inicio')

    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'crear_metodo':
            nombre = request.POST.get('nombremetodopago')
            descripcion = request.POST.get('descripcionmetodopago')
            numero_cuenta = request.POST.get('urlmetodopago')
            
            if nombre and numero_cuenta:
                MetodoPago.objects.create(
                    nombremetodopago=nombre,
                    descripcionmetodopago=descripcion,
                    urlmetodopago=numero_cuenta,
                    estadometodopago='Activo' # Guardado como texto según tu BD
                )
                messages.success(request, "¡Nuevo método de pago agregado exitosamente a la cooperativa!")
            else:
                messages.error(request, "Faltan datos obligatorios para registrar el método.")
        
        elif action == 'eliminar_metodo':
            id_metodo = request.POST.get('id_metodo')
            metodo = MetodoPago.objects.filter(idmetodopago=id_metodo).first()
            if metodo:
                metodo.delete()
                messages.success(request, "El método de pago fue eliminado de forma segura.")
        
    # Siempre redirigimos de vuelta a la central de mando
    return redirect('dashboard')
# ===============================================================================================================================================

# ===============================================================================================================================================
# 5. NEGOCIO (Logica aparte de la financiera)
# ===============================================================================================================================================
@login_required
def venta_cartones(request):
    jugador = Jugador.objects.filter(cedulaidentidadjugador=request.user.username).first()
    if not jugador:
        messages.warning(request, "Debes activar tu perfil de juego para entrar a la tienda.")
        return redirect('registro_jugador')

    if jugador.estadocuentajugador != 'Activo':
        messages.error(request, "Tu cuenta de jugador se encuentra suspendida o inactiva. No puedes realizar compras.")
        return redirect('inicio')

    if request.method == 'POST':
        id_bingo = request.POST.get('id_bingo')
        bingo = get_object_or_404(Bingo, idbingo=id_bingo)
        cartones_catalogo_ids = request.POST.getlist('cartones_catalogo')
        cartones_generados_json = request.POST.get('cartones_generados', '[]')
        
        try: cartones_generados = json.loads(cartones_generados_json)
        except Exception: cartones_generados = []

        cantidad_total_compra = len(cartones_catalogo_ids) + len(cartones_generados)

        if cantidad_total_compra == 0:
            messages.error(request, "No seleccionaste ni generaste ningún cartón para comprar.")
            return redirect('venta_cartones')

        cartones_ya_comprados = CartonPartidaBingo.objects.filter(idjugador=jugador, idpartida__idbingo=bingo).values('idcarton').distinct().count()

        # =========================================================
        # NUEVA VALIDACIÓN: BLOQUEO ESTRICTO DE 50 CARTONES GLOBALES
        # =========================================================
        if (cartones_ya_comprados + cantidad_total_compra) > 50:
            disponibles = 50 - cartones_ya_comprados
            if disponibles > 0:
                mensaje = f"Límite excedido. Ya posees {cartones_ya_comprados} cartones. Solo puedes adquirir {disponibles} cartones más para este evento."
            else:
                mensaje = "Has alcanzado el límite máximo de 50 cartones para este evento. No puedes realizar más compras."
            
            messages.error(request, mensaje)
            return redirect('venta_cartones')
        # =========================================================
        

        precio_unitario = bingo.preciocarton
        total_pagar = precio_unitario * cantidad_total_compra

        # 1. IDENTIFICAR TIPO DE MONEDA DE VENTA
        tipo_moneda_venta = bingo.idunidad_venta.tipomoneda

        # 2. VALIDAR FONDOS CORRESPONDIENTES
        if tipo_moneda_venta == 'Efectivo':
            if jugador.saldocreditojugador < total_pagar:
                messages.error(request, f"Fondos insuficientes. El total es ${total_pagar} y dispones de ${jugador.saldocreditojugador} (Saldo Real).")
                return redirect('venta_cartones')
        else:
            if jugador.saldovirtualjugador < total_pagar:
                messages.error(request, f"Fondos insuficientes. El total es {total_pagar} y dispones de {jugador.saldovirtualjugador} (Saldo Virtual).")
                return redirect('venta_cartones')

        partidas = PartidaBingo.objects.filter(idbingo=bingo)
        cartones_a_asignar = []

        if cartones_catalogo_ids:
            usados = CartonPartidaBingo.objects.filter(idpartida__in=partidas, idcarton__in=cartones_catalogo_ids).exists()
            if usados:
                messages.error(request, "Oops. Un jugador más rápido compró uno de los cartones de catálogo que elegiste. Vuelve a intentarlo.")
                return redirect('venta_cartones')
            catalogo_validos = Carton.objects.filter(idcarton__in=cartones_catalogo_ids)
            cartones_a_asignar.extend(list(catalogo_validos))

        if cartones_generados:
            nuevos_cartones_db = [Carton(codigocarton=c_data['codigo'], matriznumeros=c_data['matriz'], esmaestro=False) for c_data in cartones_generados]
            Carton.objects.bulk_create(nuevos_cartones_db)
            codigos_creados = [c['codigo'] for c in cartones_generados]
            cartones_temporales = Carton.objects.filter(codigocarton__in=codigos_creados)
            cartones_a_asignar.extend(list(cartones_temporales))

        try:
            if tipo_moneda_venta == 'Efectivo':
                jugador.saldocreditojugador -= total_pagar
            else:
                jugador.saldovirtualjugador -= total_pagar
            jugador.save()
            
            nuevas_asignaciones = []
            for carton in cartones_a_asignar:
                for partida in partidas:
                    nuevas_asignaciones.append(CartonPartidaBingo(idjugador=jugador, idpartida=partida, idcarton=carton, preciopagado=precio_unitario, estadocarton='Vendido', fechacompra=datetime.now()))
            
            if nuevas_asignaciones:
                CartonPartidaBingo.objects.bulk_create(nuevas_asignaciones)
            
            # ==========================================
            # MAGIA 5: AVISAR A LA TIENDA EN TIEMPO REAL
            # ==========================================
            channel_layer = get_channel_layer()
            for carton in cartones_a_asignar:
                # El grupo de la tienda usa el ID del Bingo maestro
                async_to_sync(channel_layer.group_send)(
                    f'bingo_tienda_{bingo.idbingo}',
                    {
                        'type': 'evento_tienda',
                        'datos': {
                            'evento': 'carton_vendido',
                            'id_carton': carton.idcarton
                        }
                    }
                )
            # ==========================================
            
            messages.success(request, f"¡Adrenalina pura! Tus {cantidad_total_compra} cartones han sido registrados en la base de datos para el evento '{bingo.titulobingo}'.")
            return redirect('venta_cartones')
        
        except Exception as e:
            messages.error(request, f"Fallo crítico en la transacción: {str(e)}")
            return redirect('venta_cartones')

    bingos_disponibles = Bingo.objects.exclude(estadobingo__in=['Finalizado', 'Cancelado']).filter(partidabingo__isnull=False).distinct()
    bingos_data = []
    for b in bingos_disponibles:
        comprados = CartonPartidaBingo.objects.filter(idjugador=jugador, idpartida__idbingo=b).values('idcarton').distinct().count()
        porcentaje_barra = min(int((comprados / 15) * 100), 100)
        usados_ids = CartonPartidaBingo.objects.filter(idpartida__idbingo=b).values_list('idcarton', flat=True)
        catalogo = Carton.objects.filter(esmaestro=True).exclude(idcarton__in=usados_ids)[:12]

        # =========================================================
        # CÁLCULO DEL POZO DINÁMICO (SOPORTE MULTIDIVISA - 45%)
        # =========================================================
        vendidos = CartonPartidaBingo.objects.filter(idpartida__idbingo=b).values('idcarton').distinct().count()
        
        # 1. Extraemos las tasas de conversión
        tasa_venta = float(b.idunidad_venta.tasaconversionmoneda)
        tasa_premio = float(b.idunidad_premio.tasaconversionmoneda)
        
        # 2. Convertimos TODO a Dólares
        ingreso_en_dolares = float(vendidos * b.preciocarton) * tasa_venta
        
        # 3. El 45% va directo al Pozo Mayor
        fondo_pozo_dolares = ingreso_en_dolares * 0.45
        premio_base_dolares = float(b.premiomayor) * tasa_premio
        
        # 4. Comparamos en Dólares y lo transformamos de vuelta
        if fondo_pozo_dolares > premio_base_dolares:
            b.pozo_dinamico_actual = fondo_pozo_dolares / tasa_premio
        else:
            b.pozo_dinamico_actual = float(b.premiomayor)
        # =========================================================

        bingos_data.append({'bingo': b, 'comprados': comprados, 'porcentaje': porcentaje_barra, 'catalogo': catalogo})

    contexto = {'jugador': jugador, 'bingos_data': bingos_data}
    return render(request, 'negocio/venta_cartones.html', contexto)

@login_required
def tienda_recargas(request):
    jugador = Jugador.objects.filter(cedulaidentidadjugador=request.user.username).first()
    if not jugador:
        messages.warning(request, "Debes activar tu perfil de juego para acceder a la billetera.")
        return redirect('registro_jugador')
        
    unidad_venta = UnidadMonetaria.objects.filter(estadomoneda=True, tipomoneda='Efectivo').first()
    unidad_virtual = UnidadMonetaria.objects.filter(estadomoneda=True, tipomoneda='Virtual').first()
    
    # Traemos las tarjetas desde la Base de Datos
    tarjetas_efectivo = TarjetaRecarga.objects.filter(estado='Activa', tiposaldo='Efectivo').order_by('preciodetarjetarecarga')
    tarjetas_virtuales = TarjetaRecarga.objects.filter(estado='Activa', tiposaldo='Virtual').order_by('preciodetarjetarecarga')
    
    # Formateamos la descripción en listas separadas por coma para el HTML
    for t in tarjetas_efectivo:
        t.lista_beneficios = [b.strip() for b in t.descripciontarjetarecarga.split(',') if b.strip()]
    for t in tarjetas_virtuales:
        t.lista_beneficios = [b.strip() for b in t.descripciontarjetarecarga.split(',') if b.strip()]
    
    contexto = {
        'jugador': jugador,
        'unidad_venta': unidad_venta,
        'unidad_virtual': unidad_virtual,
        'tarjetas_efectivo': tarjetas_efectivo,
        'tarjetas_virtuales': tarjetas_virtuales
    }
    return render(request, 'negocio/tienda_recargas.html', contexto)

@login_required
def api_catalogo_disponible(request, id_bingo):
    """Devuelve hasta 10 cartones maestros disponibles para un bingo específico"""
    if request.method == 'GET':
        usados_ids = CartonPartidaBingo.objects.filter(
            idpartida__idbingo=id_bingo
        ).values_list('idcarton', flat=True)
        
        catalogo = Carton.objects.filter(
            esmaestro=True
        ).exclude(idcarton__in=usados_ids)[:10]
        
        data = []
        for c in catalogo:
            # Reutilizamos tu lógica de parseo seguro
            matriz = c.matriznumeros
            if isinstance(matriz, str):
                try: matriz = json.loads(matriz.replace("'", '"'))
                except: continue
                
            data.append({
                'id': c.idcarton,
                'codigo': c.codigocarton,
                'matriz': matriz
            })
            
        return JsonResponse({'status': 'ok', 'catalogo': data})
# ===============================================================================================================================================

# ===============================================================================================================================================
# 6. PARTIDA (Logica y funcionamiento de las partidas)
# ===============================================================================================================================================
def estado_partida_json(request, id_partida):
    """
    Endpoint API público/interno para que los jugadores consulten el estado 
    de la ronda en tiempo real sin recargar la página.
    """
    partida = get_object_or_404(PartidaBingo, idpartidabingo=id_partida)
    
    return JsonResponse({
        'estado': partida.estadopartida,
        'hay_desempate': partida.haydesempate,
        'ganador': partida.idjugadororganador.aliasjugador if partida.idjugadororganador else None,
        'premio_efectivo': str(partida.valorpremio)
    })

@login_required
def sesion_juego(request, id_partida):
    jugador = Jugador.objects.filter(cedulaidentidadjugador=request.user.username).first()
    if not jugador:
        return redirect('registro_jugador')
        
    partida = get_object_or_404(PartidaBingo, idpartidabingo=id_partida)
    
    # 1. Obtener o crear la Plataforma Base para la Web del sistema
    plataforma, _ = PlataformaJuego.objects.get_or_create(
        nombreplataforma='Web Oficial',
        defaults={
            'urlplataforma': request.build_absolute_uri('/'),
            'descripcionplataforma': 'Acceso nativo desde la aplicación web.',
            'estadoplataforma': True
        }
    )
    
    # 2. Extraer metadatos del Navegador y Dispositivo
    user_agent = request.META.get('HTTP_USER_AGENT', 'Desconocido')
    
    # Análisis simple del dispositivo
    dispositivo = 'PC / Escritorio'
    if 'Mobile' in user_agent or 'Android' in user_agent or 'iPhone' in user_agent:
        dispositivo = 'Dispositivo Móvil'
    elif 'iPad' in user_agent or 'Tablet' in user_agent:
        dispositivo = 'Tablet'

    # Limpieza simple del nombre del navegador
    navegador = 'Otro Navegador'
    if 'Chrome' in user_agent: navegador = 'Google Chrome'
    elif 'Safari' in user_agent and 'Chrome' not in user_agent: navegador = 'Apple Safari'
    elif 'Firefox' in user_agent: navegador = 'Mozilla Firefox'
    elif 'Edge' in user_agent: navegador = 'Microsoft Edge'

    # 3. Registrar la Sesión de Juego en la Base de Datos
    # Usamos transacciones seguras para evitar duplicados críticos
    with transaction.atomic():
        # Cerramos posibles sesiones previas 'Activas' de este jugador en esta ronda
        SesionJuego.objects.filter(
            idjugador=jugador, 
            idpartida=partida, 
            estadosesion='Activa'
        ).update(estadosesion='Finalizada', fechafinsesion=timezone.now(), motivocierre='Nueva conexión establecida')

        # Creamos la nueva sesión oficial
        sesion = SesionJuego.objects.create(
            idplataforma=plataforma,
            idjugador=jugador,
            idpartida=partida,
            fechainiciosesion=timezone.now(),
            ipconexion=obtener_ip_cliente(request),
            dispositivoconexion=dispositivo,
            estadosesion='Activa',
            navegadorweb=navegador,
            tokenconexion=str(uuid.uuid4()) # Token criptográfico único
        )

    contexto = {
        'partida': partida,
        'sesion': sesion
    }
    return render(request, 'partida/sesion_juego.html', contexto)

def sala_espera(request, id_partida):
    # 1. EL GUARDIA: Interceptor de invitados
    if not request.user.is_authenticated:
        # Si no tiene cuenta, lo mandamos a tu nuevo pantallazo de bloqueo
        return render(request, 'cuentas/acceso_denegado.html')

    # 2. LÓGICA NORMAL: El resto de tu código queda intacto
    jugador = Jugador.objects.filter(cedulaidentidadjugador=request.user.username).first()
    if not jugador:
        return redirect('registro_jugador')

    # Mantengo tu búsqueda exacta con idpartidabingo
    partida = get_object_or_404(PartidaBingo, idpartidabingo=id_partida)
    
    # Si la partida ya empezó, los metemos directo al tablero
    if partida.estadopartida == 'En Juego':
        return redirect('tablero_tiempo_real', id_partida=partida.idpartidabingo)
    
    # =========================================================
    # EL PORTERO VIP: REDIRECCIÓN ABSOLUTA DESDE EL SERVIDOR
    # =========================================================
    if partida.estadopartida in ['Verificando', 'Desempate'] and partida.idbingadores:
        ids_vip = [int(i.strip()) for i in str(partida.idbingadores).split(',') if i.strip()]
        if jugador.idjugador in ids_vip:
            return redirect('sala_espera_desempate', id_partida=partida.idpartidabingo)
    # =========================================================
    
    # =========================================================
    # FIX ANTI-FANTASMAS: SOLO JUGADORES CON CONEXIÓN ACTIVA
    # =========================================================
    jugadores_en_sala = Jugador.objects.filter(
        sesionjuego__idpartida=partida,
        sesionjuego__estadosesion='Activa'
    ).distinct().order_by('aliasjugador')
    # =========================================================
        
    mensajes_historial = MensajeChat.objects.filter(idbingo=partida.idbingo).order_by('fechahora')
    
    contexto = {
        'partida': partida,
        'jugador': jugador,
        'jugadores_en_sala': jugadores_en_sala,
        'mensajes_historial': mensajes_historial # Añadir esta línea
    }
    return render(request, 'partida/sala_espera.html', contexto)

def obtener_ip_cliente(request):
    """Función helper para extraer la dirección IP real del jugador"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip

@login_required
def sala_espera_desempate(request, id_partida):
    partida = get_object_or_404(PartidaBingo, idpartidabingo=id_partida)
    
    if partida.estadopartida == 'Finalizada':
        return redirect('inicio')
    
    jugadores_en_sala = Jugador.objects.filter(
        sesionjuego__idpartida=partida,
        sesionjuego__estadosesion='Activa'
    ).distinct().order_by('aliasjugador')

    jugador = Jugador.objects.filter(cedulaidentidadjugador=request.user.username).first()
    mensajes_historial = MensajeChat.objects.filter(idbingo=partida.idbingo).order_by('fechahora')

    # NUEVO: Traer a los candidatos de la BD por si el jugador llega tarde o recarga la página
    candidatos_ids = []
    if partida.idbingadores and str(partida.idbingadores).strip().lower() != 'none':
        for id_str in str(partida.idbingadores).split(','):
            id_limpio = id_str.strip()
            if id_limpio.isdigit():
                candidatos_ids.append(int(id_limpio))
                
    candidatos = Jugador.objects.filter(idjugador__in=candidatos_ids)

    contexto = {
        'partida': partida,
        'jugador': jugador,
        'jugadores_en_sala': jugadores_en_sala,
        'mensajes_historial': mensajes_historial,
        'candidatos': candidatos # <-- ¡Enviamos los candidatos al HTML!
    }
    return render(request, 'partida/sala_espera_desempate.html', contexto)

@login_required 
def tablero_tiempo_real(request, id_partida):
    jugador = Jugador.objects.filter(cedulaidentidadjugador=request.user.username).first()
    if not jugador:
        messages.warning(request, "Necesitas un perfil de jugador para entrar a la sala.")
        return redirect('inicio')

    partida = get_object_or_404(PartidaBingo, idpartidabingo=id_partida)

    # 1. Seguridad: Redirección automática si el admin pausa o el juego termina
    if partida.estadopartida in ['Verificando', 'Desempate']:
        return redirect('sala_espera_desempate', id_partida=partida.idpartidabingo)
    elif partida.estadopartida == 'Finalizada':
        messages.info(request, "Esta ronda ha finalizado.")
        return redirect('inicio')

    # 2. Traer los cartones que ESTE jugador compró para ESTA partida
    cartones_asignados = CartonPartidaBingo.objects.filter(
        idjugador=jugador,
        idpartida=partida
    ).select_related('idcarton')

    # 3. Procesar las bolas cantadas para saber cuáles colorear
    bolas_str = partida.bolascantadas.replace('B','').replace('I','').replace('N','').replace('G','').replace('O','')
    bolas_llamadas = [int(b.strip()) for b in bolas_str.split(',') if b.strip().isdigit()]

    # 4. Preparar las matrices para que HTML las dibuje fácilmente (fila por fila)
    for asignacion in cartones_asignados:
        matriz = asignacion.idcarton.matriznumeros
        if isinstance(matriz, str):
            matriz = json.loads(matriz.replace("'", '"'))
        
        filas = []
        for i in range(5):
            fila = [
                matriz['B'][i], matriz['I'][i], matriz['N'][i], matriz['G'][i], matriz['O'][i]
            ]
            filas.append(fila)
        asignacion.filas_matriz = filas

    # 5. Obtener la lista de todos los jugadores únicos en esta partida
    jugadores_en_sala = Jugador.objects.filter(
        sesionjuego__idpartida=partida,
        sesionjuego__estadosesion='Activa'
    ).distinct().order_by('aliasjugador')

    mensajes_historial = MensajeChat.objects.filter(idbingo=partida.idbingo).order_by('fechahora')

    contexto = {
        'partida': partida,
        'jugador': jugador,
        'cartones_asignados': cartones_asignados,
        'bolas_llamadas': bolas_llamadas,
        'jugadores_en_sala': jugadores_en_sala,
        'mensajes_historial': mensajes_historial, # Añadir esta línea
    }
    return render(request, 'partida/tablero_tiempo_real.html', contexto)

@login_required
def tablero_admin(request, id_partida):
    if not request.user.is_staff:
        messages.error(request, "Acceso denegado. Zona exclusiva de administración.")
        return redirect('inicio')
        
    partida = get_object_or_404(PartidaBingo, idpartidabingo=id_partida)
    
    # =======================================================
    # NUEVO: BOTÓN MANUAL DE INICIO DE PARTIDA Y FIX DE ESTADO
    # =======================================================
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'iniciar_partida' and partida.estadopartida == 'Programada':
            # 1. Actualizamos la ronda actual
            partida.estadopartida = 'En Juego'
            partida.horainiciopartida = timezone.now()
            partida.save()
            
            # 2. FIX: ACTUALIZAR EL BINGO PADRE A 'EN CURSO'
            bingo_padre = partida.idbingo
            if bingo_padre.estadobingo == 'Programado':
                bingo_padre.estadobingo = 'En Curso'
                bingo_padre.save()
            
            # 3. El Árbitro sopla el silbato por WebSockets
            channel_layer = get_channel_layer()
            async_to_sync(channel_layer.group_send)(
                f'bingo_partida_{partida.idpartidabingo}',
                {'type': 'evento_partida', 'datos': {'evento': 'estado_cambiado', 'nuevo_estado': 'En Juego'}}
            )
            messages.success(request, f"¡Pitazo inicial! La ronda ha comenzado y el Bingo está En Curso.")
            return redirect('tablero_admin', id_partida=partida.idpartidabingo)
    # =======================================================

    # 1. Limpiar y obtener las bolas cantadas como una lista de enteros
    bolas_str = partida.bolascantadas.replace('B','').replace('I','').replace('N','').replace('G','').replace('O','')
    bolas_llamadas = [int(b.strip()) for b in bolas_str.split(',') if b.strip().isdigit()]

    # 2. Construir la estructura del Tablero Maestro (1 al 75)
    tablero_maestro = {
        'B': {'rango': range(1, 16), 'color': 'primary'},     
        'I': {'rango': range(16, 31), 'color': 'danger'},     
        'N': {'rango': range(31, 46), 'color': 'secondary'},  
        'G': {'rango': range(46, 61), 'color': 'success'},    
        'O': {'rango': range(61, 76), 'color': 'warning'}     
    }

    # NUEVO: Obtener la lista de todos los jugadores únicos en esta partida para el radar
    jugadores_en_sala = Jugador.objects.filter(
        sesionjuego__idpartida=partida,
        sesionjuego__estadosesion='Activa'
    ).distinct().order_by('aliasjugador')

    contexto = {
        'partida': partida,
        'bolas_llamadas': bolas_llamadas,
        'tablero_maestro': tablero_maestro,
        'jugadores_en_sala': jugadores_en_sala, # <--- Enviamos los jugadores al Tablero Admin
    }
    return render(request, 'partida/tablero_admin.html', contexto)

@login_required
def sacar_bola_api(request, id_partida):
    if not request.method == 'POST' or not request.user.is_staff:
        return JsonResponse({'error': 'Acceso denegado'}, status=403)

    partida = get_object_or_404(PartidaBingo, idpartidabingo=id_partida)
    
    if partida.estadopartida != 'En Juego':
        return JsonResponse({'error': 'La partida no está en curso'}, status=400)

    bolas_str = partida.bolascantadas.replace('B','').replace('I','').replace('N','').replace('G','').replace('O','')
    bolas_llamadas = [int(b.strip()) for b in bolas_str.split(',') if b.strip().isdigit()]

    # --- NUEVO: CAPTURAR EL NÚMERO TRAMPA DEL BODY ---
    try:
        cuerpo_peticion = json.loads(request.body)
        numero_ninja = int(cuerpo_peticion.get('numero_forzado', 0))
    except Exception:
        numero_ninja = 0
    # -------------------------------------------------

    bolas_disponibles = [i for i in range(1, 76) if i not in bolas_llamadas]
    if not bolas_disponibles:
        return JsonResponse({'error': 'No hay más bolas disponibles'}, status=400)

    # Si llegó el número trampa y no ha salido, lo forzamos. Si no, usamos Random clásico
    if numero_ninja > 0 and numero_ninja in bolas_disponibles:
        nueva_bola = numero_ninja
    else:
        nueva_bola = random.choice(bolas_disponibles)

    bolas_llamadas.append(nueva_bola)

    partida.ultimabola = nueva_bola
    partida.bolascantadas = ",".join(map(str, bolas_llamadas))
    partida.save()

    channel_layer = get_channel_layer()
    async_to_sync(channel_layer.group_send)(
        f'bingo_partida_{id_partida}',
        {
            'type': 'evento_partida',
            'datos': {
                'evento': 'nueva_bola',
                'numero': nueva_bola
            }
        }
    )

    return JsonResponse({'status': 'ok', 'bola_extraida': nueva_bola})

@login_required
def desempate_admin(request, id_partida):
    if not request.user.is_staff:
        return redirect('inicio')
        
    partida = get_object_or_404(PartidaBingo, idpartidabingo=id_partida)
    channel_layer = get_channel_layer() 
    
    if partida.estadopartida == 'En Juego':
        partida.estadopartida = 'Verificando'
        partida.save()
        
        async_to_sync(channel_layer.group_send)(
            f'bingo_partida_{id_partida}',
            {'type': 'evento_partida', 'datos': {'evento': 'estado_cambiado', 'nuevo_estado': 'Verificando'}}
        )

    if request.method == 'POST':
        decision = request.POST.get('decision_desempate')
        
        if decision == 'si':
            partida.estadopartida = 'Desempate'
            partida.haydesempate = True
            partida.save()
            
            async_to_sync(channel_layer.group_send)(
                f'bingo_partida_{id_partida}',
                {'type': 'evento_partida', 'datos': {'evento': 'estado_cambiado', 'nuevo_estado': 'Desempate'}}
            )
            
            messages.info(request, "Modo Desempate Activado. Prepare la consola.")
            return redirect('consola_juego', id_partida=partida.idpartidabingo)
            
        elif decision == 'no':
            codigo_ganador = request.POST.get('codigo_ganador_unico')
            resultado = validar_carton_hibrido(codigo_ganador, partida.idpartidabingo)
                                                                                                
            if resultado['existe'] and resultado['valido']:
                partida.estadopartida = 'Finalizada'
                partida.idjugadororganador_id = resultado['id_jugador']
                partida.horafin = timezone.now() 
                partida.save()
                
                # ==========================================
                # MAGIA FINANCIERA: PAGO AUTOMÁTICO DE PREMIOS
                # ==========================================
                es_pozo_mayor = (partida.premiomaterial == '[POZO_MAYOR]')
                monto_a_pagar = partida.idbingo.premiomayor if es_pozo_mayor else partida.valorpremio
                
                if monto_a_pagar and monto_a_pagar > 0:
                    jugador_ganador = Jugador.objects.get(idjugador=resultado['id_jugador'])
                    tipo_moneda = partida.idbingo.idunidad_premio.tipomoneda
                    if tipo_moneda == 'Efectivo':
                        jugador_ganador.saldocreditojugador += monto_a_pagar
                    else:
                        jugador_ganador.saldovirtualjugador += monto_a_pagar
                    jugador_ganador.save()
                
                # Logística del Premio Físico (Si NO es el pozo mayor)
                if not es_pozo_mayor and partida.premiomaterial and partida.premiomaterial != 'Ninguno':
                    partida.estadopremiomaterial = 'Pendiente'
                    
                partida.save()

                # ==========================================
                # ÁRBITRO DIGITAL: RELEVO Y ENRUTAMIENTO (FASE 3)
                # ==========================================
                siguiente_partida = PartidaBingo.objects.filter(
                    idbingo=partida.idbingo,
                    idpartidabingo__gt=partida.idpartidabingo
                ).order_by('idpartidabingo').first()

                if siguiente_partida:
                    destino_admin = redirect('tablero_admin', id_partida=siguiente_partida.idpartidabingo)
                else:
                    bingo_actual = partida.idbingo
                    bingo_actual.estadobingo = 'Finalizado'
                    bingo_actual.save()
                    destino_admin = redirect('dashboard')
                # ==========================================
                
                id_siguiente = siguiente_partida.idpartidabingo if siguiente_partida else None
                
                async_to_sync(channel_layer.group_send)(
                    f'bingo_partida_{id_partida}',
                    {'type': 'evento_partida', 'datos': {
                        'evento': 'estado_cambiado', 
                        'nuevo_estado': 'Finalizada',
                        'ganador': resultado['jugador'],
                        'id_siguiente_partida': id_siguiente
                    }}
                )
                
                messages.success(request, f"¡Partida finalizada! Ganador único asignado: {resultado['jugador']}")
                return destino_admin
            else:
                messages.error(request, "El código ingresado no es válido o no completó el cartón.")
                return redirect('desempate_admin', id_partida=partida.idpartidabingo)
        
        elif decision == 'reanudar':
            partida.estadopartida = 'En Juego'
            partida.save()
            
            async_to_sync(channel_layer.group_send)(
                f'bingo_partida_{id_partida}',
                {'type': 'evento_partida', 'datos': {'evento': 'estado_cambiado', 'nuevo_estado': 'En Juego'}}
            )
            
            messages.success(request, "Falsa alarma. La partida ha sido reanudada.")
            return redirect('tablero_admin', id_partida=partida.idpartidabingo)

    # =========================================================
    # ESCÁNER DE GANADORES WEB EN TIEMPO REAL (RADAR ESTRICTO)
    # =========================================================
    # 2. Diccionario Maestro de Patrones (Limpiado y Optimizado)
    patrones = {
        'Tabla Llena': [[0,1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24]],
        'Las Cuatro Esquinas': [[0, 4, 20, 24]],
        'En Diagonal': [[0, 6, 12, 18, 24], [4, 8, 12, 16, 20]],
        'Forma de X': [[0, 4, 6, 8, 12, 16, 18, 20, 24]],
        'Forma de Cruz': [[2, 7, 10, 11, 12, 13, 14, 17, 22]],
        'Marco de Foto': [[0,1,2,3,4, 5,9, 10,14, 15,19, 20,21,22,23,24]],
        'Linea Vertical': [
            [0, 5, 10, 15, 20], [1, 6, 11, 16, 21], [2, 7, 12, 17, 22], [3, 8, 13, 18, 23], [4, 9, 14, 19, 24]
        ],
        'Forma de L': [
            [0, 5, 10, 15, 20, 21, 22, 23, 24], 
            [0, 1, 2, 3, 4, 5, 10, 15, 20],     
            [0, 1, 2, 3, 4, 9, 14, 19, 24],     
            [20, 21, 22, 23, 24, 19, 14, 9, 4]  
        ],
        'Forma de C': [
            [0, 1, 2, 3, 4, 5, 10, 15, 20, 21, 22, 23, 24], 
            [0, 1, 2, 3, 4, 9, 14, 19, 24, 20, 21, 22, 23], 
            [0, 5, 10, 15, 20, 21, 22, 23, 24, 4, 9, 14, 19],
            [20, 15, 10, 5, 0, 1, 2, 3, 4, 9, 14, 19, 24]
        ],
        'Forma de U': [
            [0, 5, 10, 15, 20, 21, 22, 23, 24, 4, 9, 14, 19],
            [20, 15, 10, 5, 0, 1, 2, 3, 4, 9, 14, 19, 24], 
            [0, 1, 2, 3, 4, 5, 10, 15, 20, 21, 22, 23, 24], 
            [0, 1, 2, 3, 4, 9, 14, 19, 24, 20, 21, 22, 23]  
        ],
        'Forma de T': [
            [0,1,2,3,4, 7, 12, 17, 22], 
            [20,21,22,23,24, 17, 12, 7, 2], 
            [4,9,14,19,24, 13, 12, 11, 10], 
            [0,5,10,15,20, 11, 12, 13, 14]  
        ],
        'Forma de H': [
            [0, 5, 10, 15, 20, 11, 12, 13, 4, 9, 14, 19, 24], 
            [0, 1, 2, 3, 4, 7, 12, 17, 20, 21, 22, 23, 24]       
        ],
        'Forma de Z': [
            [0,1,2,3,4, 8, 12, 16, 20,21,22,23,24], 
            [0,1,2,3,4, 6, 12, 18, 20,21,22,23,24], 
            [4,9,14,19,24, 18, 12, 6, 0,5,10,15,20], 
            [0,5,10,15,20, 16, 12, 8, 4,9,14,19,24]  
        ],
        'Forma de Flecha': [
            [2, 6, 8, 12, 17, 22],   
            [22, 16, 18, 12, 7, 2],  
            [10, 6, 16, 12, 13, 14], 
            [14, 8, 18, 12, 11, 10]  
        ]
    }
    
    # FIX: Búsqueda Insensible a Mayúsculas y Espacios para el Radar Web
    modalidad_limpia = str(partida.modalidad_victoria).strip().lower()
    patrones_lower = {k.lower(): v for k, v in patrones.items()}
    marcadas_requeridas = patrones_lower.get(modalidad_limpia, patrones_lower['tabla llena'])
    
    cartones_en_juego = CartonPartidaBingo.objects.filter(
        idpartida=partida,
    ).select_related('idcarton', 'idjugador')
    
    ganadores_web = []
    jugadores_en_radar = set() # Control para evitar clones en pantalla

    for c in cartones_en_juego:
        id_jugador_actual = c.idjugador.idjugador
            
        if id_jugador_actual in jugadores_en_radar:
            continue
            
        marcados_db = []
        if getattr(c, 'numerosmarcados', None):
            try: marcados_db = json.loads(c.numerosmarcados)
            except: 
                try: marcados_db = ast.literal_eval(c.numerosmarcados)
                except: pass
        
        if not marcados_db: continue
        
        # FIX: Limpiamos cada número clickeado para evitar espacios fantasma
        marcados_str = [str(num).strip() for num in marcados_db]
        
        matriz = c.idcarton.matriznumeros
        if isinstance(matriz, str):
            try: matriz = json.loads(matriz.replace("'", '"'))
            except: continue
            
        celdas = []
        for i in range(5):
            celdas.extend([matriz['B'][i], matriz['I'][i], matriz['N'][i], matriz['G'][i], matriz['O'][i]])
            
        # LÓGICA DE VALIDACIÓN MULTI-PATRÓN
        es_ganador_global = False
        for opcion in marcadas_requeridas:
            es_ganador_opcion = True
            for idx in opcion:
                if idx == 12: continue 
                
                # FIX: Limpiamos la celda extraída de la BD antes de compararla
                if str(celdas[idx]).strip() not in marcados_str:
                    es_ganador_opcion = False
                    break
                    
            if es_ganador_opcion:
                es_ganador_global = True
                break
                
        if es_ganador_global:
            if not c.fechaganador:
                c.fechaganador = timezone.now()
                c.save(update_fields=['fechaganador'])
                
            ganadores_web.append(c)
            jugadores_en_radar.add(id_jugador_actual)

    ganadores_web.sort(key=lambda x: x.fechaganador)

    contexto = {
        'partida': partida,
        'ganadores_web': ganadores_web
    }
    return render(request, 'partida/desempate_admin.html', contexto)

@login_required
def consola_juego(request, id_partida):
    if not request.user.is_staff:
        return redirect('inicio')
        
    partida = get_object_or_404(PartidaBingo, idpartidabingo=id_partida)
    
    if partida.estadopartida == 'Finalizada':
        messages.info(request, "Esta partida ya ha finalizado.")
        return redirect('dashboard')
        
    # =========================================================
    # NUEVO: CÁLCULO DEL POZO DINÁMICO PARA EL DESEMPATE
    # =========================================================
    vendidos = CartonPartidaBingo.objects.filter(idpartida__idbingo=partida.idbingo).values('idcarton').distinct().count()
    tasa_venta = float(partida.idbingo.idunidad_venta.tasaconversionmoneda)
    tasa_premio = float(partida.idbingo.idunidad_premio.tasaconversionmoneda)
    ingreso_en_dolares = float(vendidos * partida.idbingo.preciocarton) * tasa_venta
    fondo_pozo_dolares = ingreso_en_dolares * 0.45
    premio_base_dolares = float(partida.idbingo.premiomayor) * tasa_premio
    
    if fondo_pozo_dolares > premio_base_dolares:
        partida.idbingo.pozo_dinamico_actual = fondo_pozo_dolares / tasa_premio
    else:
        partida.idbingo.pozo_dinamico_actual = float(partida.idbingo.premiomayor)
    # =========================================================
    
    # 1. OBTENER CANDIDATOS (Forma 100% segura y blindada)
    candidatos_ids = []
    if partida.idbingadores and str(partida.idbingadores).strip().lower() != 'none':
        for id_str in str(partida.idbingadores).split(','):
            id_limpio = id_str.strip()
            if id_limpio.isdigit(): # Solo agregamos si es un número válido
                candidatos_ids.append(int(id_limpio))
                
    candidatos = Jugador.objects.filter(idjugador__in=candidatos_ids)
    
    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'agregar_candidato':
            codigo = request.POST.get('codigo_carton')
            resultado = validar_carton_hibrido(codigo, partida.idpartidabingo)
            
            if resultado['existe'] and resultado['valido']:
                nuevo_id = str(resultado['id_jugador'])
                
                # FIX BUG: Concatenación estricta de IDs para evitar pérdidas
                ids_actuales = []
                if partida.idbingadores and str(partida.idbingadores).strip().lower() != 'none':
                    for i in str(partida.idbingadores).split(','):
                        i_limpio = i.strip()
                        if i_limpio.isdigit():
                            ids_actuales.append(i_limpio)
                            
                if nuevo_id not in ids_actuales:
                    ids_actuales.append(nuevo_id)
                    partida.idbingadores = ",".join(ids_actuales)
                    partida.save()
                    
                    channel_layer = get_channel_layer()
                    async_to_sync(channel_layer.group_send)(
                        f'bingo_partida_{id_partida}',
                        {'type': 'evento_partida', 'datos': {
                            'evento': 'invitacion_vip', 
                            'id_jugador': nuevo_id,
                            'alias': resultado['jugador'] 
                        }}
                    )
                    messages.success(request, f"¡Cartón verificado! {resultado['jugador']} agregado al desempate.")
                else:
                    messages.warning(request, "Este jugador ya está en la lista de desempate.")
            else:
                messages.error(request, "Código inválido o cartón incompleto.")
            
            return redirect('consola_juego', id_partida=id_partida)

        # ACCIÓN PARA FINALIZAR SI HAY UN ÚNICO GANADOR
        elif action == 'forzar_ganador_unico':
            ganador_id = request.POST.get('id_unico_ganador')
            if ganador_id:
                partida.idjugadororganador_id = ganador_id
                partida.estadopartida = 'Finalizada'
                partida.haydesempate = False # Corregimos el estado a Falsa Alarma de desempate
                partida.horafin = timezone.now() 
                partida.save()
                
                es_pozo_mayor = (partida.premiomaterial == '[POZO_MAYOR]')
                monto_a_pagar = partida.idbingo.premiomayor if es_pozo_mayor else partida.valorpremio
                
                if monto_a_pagar and monto_a_pagar > 0:
                    jugador_ganador = Jugador.objects.get(idjugador=ganador_id)
                    tipo_moneda = partida.idbingo.idunidad_premio.tipomoneda
                    if tipo_moneda == 'Efectivo':
                        jugador_ganador.saldocreditojugador += monto_a_pagar
                    else:
                        jugador_ganador.saldovirtualjugador += monto_a_pagar
                    jugador_ganador.save()
                    
                if not es_pozo_mayor and partida.premiomaterial and partida.premiomaterial != 'Ninguno':
                    partida.estadopremiomaterial = 'Pendiente'
                partida.save()
                
                siguiente_partida = PartidaBingo.objects.filter(
                    idbingo=partida.idbingo, idpartidabingo__gt=partida.idpartidabingo
                ).order_by('idpartidabingo').first()

                if siguiente_partida:
                    destino_admin = redirect('tablero_admin', id_partida=siguiente_partida.idpartidabingo)
                else:
                    bingo_actual = partida.idbingo
                    bingo_actual.estadobingo = 'Finalizado'
                    bingo_actual.save()
                    destino_admin = redirect('dashboard')
                
                ganador_obj = Jugador.objects.get(idjugador=ganador_id)
                channel_layer = get_channel_layer()
                async_to_sync(channel_layer.group_send)(
                    f'bingo_partida_{id_partida}',
                    {'type': 'evento_partida', 'datos': {
                        'evento': 'estado_cambiado', 'nuevo_estado': 'Finalizada',
                        'ganador': ganador_obj.aliasjugador, 'id_siguiente_partida': siguiente_partida.idpartidabingo if siguiente_partida else None
                    }}
                )
                messages.success(request, "¡Ronda liquidada! Se ha asignado el premio al único ganador verificado.")
                return destino_admin
            
        # ACCIÓN PARA CANCELAR EL DESEMPATE Y REGRESAR
        elif action == 'cancelar_desempate':
            partida.estadopartida = 'En Juego'
            partida.haydesempate = False
            partida.idbingadores = "" # Limpiamos a los candidatos
            partida.sorteodesempate = {} # Limpiamos los tiros
            partida.save()
            
            # Le avisamos a todos los jugadores que volvemos al tablero
            channel_layer = get_channel_layer()
            async_to_sync(channel_layer.group_send)(
                f'bingo_partida_{id_partida}',
                {'type': 'evento_partida', 'datos': {'evento': 'estado_cambiado', 'nuevo_estado': 'En Juego'}}
            )
            
            messages.success(request, "Desempate abortado. La partida se ha reanudado.")
            return redirect('tablero_admin', id_partida=partida.idpartidabingo)
        
        elif action == 'registrar_tiro_desempate':
            id_jugador_tiro = request.POST.get('id_jugador_tiro')
            numero_tiro = int(request.POST.get('numero_tiro'))
            
            sorteo = partida.sorteodesempate or {}
            sorteo[str(id_jugador_tiro)] = numero_tiro
            partida.sorteodesempate = sorteo
            partida.save()
            
            # Disparo de animación individual
            channel_layer = get_channel_layer()
            async_to_sync(channel_layer.group_send)(
                f'bingo_partida_{id_partida}',
                {'type': 'evento_partida', 'datos': {
                    'evento': 'tiro_individual', 
                    'id_jugador': id_jugador_tiro, 
                    'numero': numero_tiro
                }}
            )

            ids_actuales = [str(i.strip()) for i in str(partida.idbingadores).split(',') if i.strip()]
            completado = all(candidato in sorteo for candidato in ids_actuales)
            
            if completado:
                ganador_id = max(sorteo, key=sorteo.get)
                ganador_numero = sorteo[ganador_id]
                ganador_obj = Jugador.objects.filter(idjugador=int(ganador_id)).first()
                ganador_nombre = ganador_obj.aliasjugador if ganador_obj else "Jugador Oficial"
                
                channel_layer = get_channel_layer()
                async_to_sync(channel_layer.group_send)(
                    f'bingo_partida_{id_partida}',
                    {
                        'type': 'evento_partida',
                        'datos': {
                            'evento': 'desempate_completado',
                            'ganador_id': ganador_id,
                            'ganador_numero': ganador_numero,
                            'ganador_nombre': ganador_nombre
                        }
                    }
                )
            
            return JsonResponse({'status': 'ok', 'completado': completado})
            
        elif action == 'resolver_desempate':
            ganador_id = request.POST.get('ganador_final')
            bola_mayor = request.POST.get('bola_mayor')
            
            if ganador_id and bola_mayor:
                partida.idjugadororganador_id = ganador_id
                partida.bolamayordesempate = bola_mayor
                partida.estadopartida = 'Finalizada'
                partida.horafin = timezone.now() 
                partida.save()
                
                es_pozo_mayor = (partida.premiomaterial == '[POZO_MAYOR]')
                monto_a_pagar = partida.idbingo.premiomayor if es_pozo_mayor else partida.valorpremio
                
                if monto_a_pagar and monto_a_pagar > 0:
                    jugador_ganador = Jugador.objects.get(idjugador=ganador_id)
                    tipo_moneda = partida.idbingo.idunidad_premio.tipomoneda
                    if tipo_moneda == 'Efectivo':
                        jugador_ganador.saldocreditojugador += monto_a_pagar
                    else:
                        jugador_ganador.saldovirtualjugador += monto_a_pagar
                    jugador_ganador.save()
                    
                if not es_pozo_mayor and partida.premiomaterial and partida.premiomaterial != 'Ninguno':
                    partida.estadopremiomaterial = 'Pendiente'

                partida.save()
                
                siguiente_partida = PartidaBingo.objects.filter(
                    idbingo=partida.idbingo,
                    idpartidabingo__gt=partida.idpartidabingo
                ).order_by('idpartidabingo').first()

                if siguiente_partida:
                    destino_admin = redirect('tablero_admin', id_partida=siguiente_partida.idpartidabingo)
                else:
                    bingo_actual = partida.idbingo
                    bingo_actual.estadobingo = 'Finalizado'
                    bingo_actual.save()
                    destino_admin = redirect('dashboard')
                
                id_siguiente = siguiente_partida.idpartidabingo if siguiente_partida else None
                
                ganador_obj = Jugador.objects.get(idjugador=ganador_id)
                channel_layer = get_channel_layer()
                async_to_sync(channel_layer.group_send)(
                    f'bingo_partida_{id_partida}',
                    {'type': 'evento_partida', 'datos': {
                        'evento': 'estado_cambiado', 
                        'nuevo_estado': 'Finalizada',
                        'ganador': ganador_obj.aliasjugador,
                        'id_siguiente_partida': id_siguiente
                    }}
                )
                
                messages.success(request, "¡Desempate resuelto! El ganador ha sido registrado y la ronda ha finalizado.")
                return destino_admin
            else:
                messages.error(request, "Debe seleccionar un ganador e ingresar la bola mayor.")
                return redirect('consola_juego', id_partida=id_partida)

    # =========================================================
    # ESCÁNER DE GANADORES WEB EN TIEMPO REAL (RADAR ESTRICTO)
    # =========================================================
    patrones = {
        'Tabla Llena': [[0,1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24]],
        'Las Cuatro Esquinas': [[0, 4, 20, 24]],
        'En Diagonal': [[0, 6, 12, 18, 24], [4, 8, 12, 16, 20]],
        'Forma de X': [[0, 4, 6, 8, 12, 16, 18, 20, 24]],
        'Forma de Cruz': [[2, 7, 10, 11, 12, 13, 14, 17, 22]],
        'Marco de Foto': [[0,1,2,3,4, 5,9, 10,14, 15,19, 20,21,22,23,24]],
        'Linea Vertical': [
            [0, 5, 10, 15, 20], [1, 6, 11, 16, 21], [2, 7, 12, 17, 22], [3, 8, 13, 18, 23], [4, 9, 14, 19, 24]
        ],
        'Forma de L': [
            [0, 5, 10, 15, 20, 21, 22, 23, 24], 
            [0, 1, 2, 3, 4, 5, 10, 15, 20],     
            [0, 1, 2, 3, 4, 9, 14, 19, 24],     
            [20, 21, 22, 23, 24, 19, 14, 9, 4]  
        ],
        'Forma de C': [
            [0, 1, 2, 3, 4, 5, 10, 15, 20, 21, 22, 23, 24], 
            [0, 1, 2, 3, 4, 9, 14, 19, 24, 20, 21, 22, 23], 
            [0, 5, 10, 15, 20, 21, 22, 23, 24, 4, 9, 14, 19],
            [20, 15, 10, 5, 0, 1, 2, 3, 4, 9, 14, 19, 24]
        ],
        'Forma de U': [
            [0, 5, 10, 15, 20, 21, 22, 23, 24, 4, 9, 14, 19],
            [20, 15, 10, 5, 0, 1, 2, 3, 4, 9, 14, 19, 24], 
            [0, 1, 2, 3, 4, 5, 10, 15, 20, 21, 22, 23, 24], 
            [0, 1, 2, 3, 4, 9, 14, 19, 24, 20, 21, 22, 23]  
        ],
        'Forma de T': [
            [0,1,2,3,4, 7, 12, 17, 22], 
            [20,21,22,23,24, 17, 12, 7, 2], 
            [4,9,14,19,24, 13, 12, 11, 10], 
            [0,5,10,15,20, 11, 12, 13, 14]  
        ],
        'Forma de H': [
            [0, 5, 10, 15, 20, 11, 12, 13, 4, 9, 14, 19, 24], 
            [0, 1, 2, 3, 4, 7, 12, 17, 20, 21, 22, 23, 24]       
        ],
        'Forma de Z': [
            [0,1,2,3,4, 8, 12, 16, 20,21,22,23,24], 
            [0,1,2,3,4, 6, 12, 18, 20,21,22,23,24], 
            [4,9,14,19,24, 18, 12, 6, 0,5,10,15,20], 
            [0,5,10,15,20, 16, 12, 8, 4,9,14,19,24]  
        ],
        'Forma de Flecha': [
            [2, 6, 8, 12, 17, 22],   
            [22, 16, 18, 12, 7, 2],  
            [10, 6, 16, 12, 13, 14], 
            [14, 8, 18, 12, 11, 10]  
        ]
    }
    
    # FIX: Búsqueda Insensible a Mayúsculas y Espacios para el Radar Web
    modalidad_limpia = str(partida.modalidad_victoria).strip().lower()
    patrones_lower = {k.lower(): v for k, v in patrones.items()}
    marcadas_requeridas = patrones_lower.get(modalidad_limpia, patrones_lower['tabla llena'])
    
    cartones_en_juego = CartonPartidaBingo.objects.filter(
        idpartida=partida,
    ).select_related('idcarton', 'idjugador')
    
    ganadores_web = []
    jugadores_en_radar = set()

    for c in cartones_en_juego:
        id_jugador_actual = c.idjugador.idjugador
        
        if 'candidatos_ids' in locals() and id_jugador_actual in candidatos_ids:
            continue
            
        if id_jugador_actual in jugadores_en_radar:
            continue
            
        marcados_db = []
        if getattr(c, 'numerosmarcados', None):
            try: marcados_db = json.loads(c.numerosmarcados)
            except: 
                try: marcados_db = ast.literal_eval(c.numerosmarcados)
                except: pass
        
        if not marcados_db: continue
            
        # FIX: Limpiamos cada número clickeado
        marcados_str = [str(num).strip() for num in marcados_db]
        
        matriz = c.idcarton.matriznumeros
        if isinstance(matriz, str):
            try: matriz = json.loads(matriz.replace("'", '"'))
            except: continue
            
        celdas = []
        for i in range(5):
            celdas.extend([matriz['B'][i], matriz['I'][i], matriz['N'][i], matriz['G'][i], matriz['O'][i]])
            
        # LÓGICA DE VALIDACIÓN MULTI-PATRÓN
        es_ganador_global = False
        for opcion in marcadas_requeridas:
            es_ganador_opcion = True
            for idx in opcion:
                if idx == 12: continue 
                
                # FIX: Limpiamos la celda extraída
                if str(celdas[idx]).strip() not in marcados_str:
                    es_ganador_opcion = False
                    break
                    
            if es_ganador_opcion:
                es_ganador_global = True
                break
                
        if es_ganador_global:
            if not c.fechaganador:
                c.fechaganador = timezone.now()
                c.save(update_fields=['fechaganador'])
                
            ganadores_web.append(c)
            jugadores_en_radar.add(id_jugador_actual)

    ganadores_web.sort(key=lambda x: x.fechaganador)

    contexto = {'partida': partida, 'candidatos': candidatos, 'ganadores_web': ganadores_web}
    return render(request, 'partida/consola_juego.html', contexto)
# ===============================================================================================================================================